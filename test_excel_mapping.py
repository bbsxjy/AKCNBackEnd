"""
Test script to verify Excel column mapping is working correctly
"""

import asyncio
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io

# Create a test Excel file with exact column headers as specified
wb = Workbook()
ws = wb.active
ws.title = "应用总追踪表"

# Set headers exactly as specified by user
headers = [
    "L2ID",
    "L2应用",
    "监管验收年份",
    "改造目标",
    "是否已完成AK",
    "是否已完成云原生",
    "当前改造阶段",
    "改造状态",
    "【计划】\n需求完成时间",
    "【计划】\n发版时间",
    "【计划】\n技术上线时间",
    "【计划】\n业务上线时间",
    "【实际】\n需求到达时间",
    "【实际】\n发版时间",
    "【实际】\n技术上线时间",
    "【实际】\n业务上线时间",
    "备注",
    "档位",
    "所属L1",
    "所属项目",
    "开发模式",
    "运维模式",
    "开发负责人",
    "开发团队",
    "运维负责人",
    "运维团队",
    "所属指标",
    "验收状态"
]

# Write headers
for i, header in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=header)

# Add sample data
sample_data = [
    ["10001", "测试应用1", 2024, "AK", "否", "否", "开发中", "研发进行中",
     "2024-03-01", "2024-03-15", "2024-04-01", "2024-04-15",
     "2024-03-05", "2024-03-20", "2024-04-10", "2024-04-20",
     "测试备注", 1, "数字金融", "AK改造项目", "敏捷", "DevOps",
     "张三", "团队A", "李四", "运维团队B", "KPI-001", "待验收"],
    ["10002", "测试应用2", 2024, "云原生", "是", "否", "测试中", "业务上线中",
     "2024-02-01", "2024-02-15", "2024-03-01", "2024-03-15",
     "2024-02-05", "2024-02-20", "2024-03-10", "2024-03-20",
     "另一个测试", 2, "零售金融", "云原生改造", "瀑布", "传统运维",
     "王五", "团队B", "赵六", "运维团队A", "KPI-002", "已验收"]
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save to bytes
excel_bytes = io.BytesIO()
wb.save(excel_bytes)
excel_content = excel_bytes.getvalue()

print(f"Created test Excel file with {len(headers)} columns and {len(sample_data)} data rows")

# Now test the import functionality
async def test_import():
    from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
    from sqlalchemy.orm import sessionmaker
    from app.services.excel_service import ExcelService
    from app.models.user import User
    from datetime import datetime, timezone

    # Create a test database connection
    DATABASE_URL = "postgresql+asyncpg://postgres:postgres@localhost:5432/akcn_db"
    engine = create_async_engine(DATABASE_URL, echo=False)
    AsyncSessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with AsyncSessionLocal() as db:
        # Create a test user or fetch existing one
        from sqlalchemy import select
        result = await db.execute(select(User).where(User.username == "admin"))
        user = result.scalar_one_or_none()

        if not user:
            user = User(
                username="admin",
                email="admin@test.com",
                department="IT",
                role="admin",
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )
            db.add(user)
            await db.commit()
            await db.refresh(user)

        # Test the import
        service = ExcelService()
        print("\n" + "="*50)
        print("Testing Excel import with correct column mapping...")
        print("="*50)

        result = await service.import_applications_from_excel(
            db=db,
            file_content=excel_content,
            user=user,
            validate_only=False  # Actually import the data
        )

        print("\n" + "="*50)
        print("Import result:")
        print(f"Success: {result.get('success')}")
        print(f"Total rows: {result.get('total_rows')}")
        print(f"Processed rows: {result.get('processed_rows')}")
        print(f"Updated rows: {result.get('updated_rows')}")
        print(f"Skipped rows: {result.get('skipped_rows')}")
        if result.get('errors'):
            print(f"Errors: {result.get('errors')}")
        print("="*50)

        # Verify the imported data
        if result.get('success'):
            from app.models.application import Application

            # Check if applications were imported with correct app_name
            for l2_id in ["10001", "10002"]:
                stmt = select(Application).where(Application.l2_id == l2_id)
                app_result = await db.execute(stmt)
                app = app_result.scalar_one_or_none()

                if app:
                    print(f"\nApplication L2ID={l2_id}:")
                    print(f"  app_name: {app.app_name}")
                    print(f"  dev_team: {app.dev_team}")
                    print(f"  dev_owner: {app.dev_owner}")
                    print(f"  current_status: {app.current_status}")
                    print(f"  overall_transformation_target: {app.overall_transformation_target}")
                else:
                    print(f"\nWARNING: Application with L2ID={l2_id} not found!")

if __name__ == "__main__":
    asyncio.run(test_import())