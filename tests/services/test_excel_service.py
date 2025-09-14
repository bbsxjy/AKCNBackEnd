"""
Unit tests for Excel service
"""

import pytest
import pandas as pd
from datetime import date, datetime
from unittest.mock import Mock, patch, AsyncMock
from openpyxl import Workbook
import io

from app.services.excel_service import ExcelService, ExcelMappingConfig, ExcelValidationError
from app.models.application import Application, ApplicationStatus, TransformationTarget
from app.models.subtask import SubTask, SubTaskStatus
from app.models.user import User, UserRole


class TestExcelService:
    """Test Excel service functionality."""

    def setup_method(self):
        """Setup test environment."""
        self.excel_service = ExcelService()
        self.config = ExcelMappingConfig()

        # Create mock user
        self.mock_user = Mock(spec=User)
        self.mock_user.id = 1
        self.mock_user.username = "test_user"
        self.mock_user.role = UserRole.MANAGER

        # Create mock database session
        self.mock_db = AsyncMock()

    def test_excel_mapping_config_initialization(self):
        """Test Excel mapping configuration."""
        config = ExcelMappingConfig()

        # Test application fields mapping
        assert 'L2 ID' in config.APPLICATION_FIELDS
        assert config.APPLICATION_FIELDS['L2 ID'] == 'l2_id'
        assert '应用名称' in config.APPLICATION_FIELDS
        assert config.APPLICATION_FIELDS['应用名称'] == 'app_name'

        # Test subtask fields mapping
        assert '应用L2 ID' in config.SUBTASK_FIELDS
        assert config.SUBTASK_FIELDS['应用L2 ID'] == 'application_l2_id'
        assert '模块名称' in config.SUBTASK_FIELDS

        # Test required fields
        assert 'l2_id' in config.APPLICATION_REQUIRED
        assert 'app_name' in config.APPLICATION_REQUIRED
        assert 'application_l2_id' in config.SUBTASK_REQUIRED

        # Test field types
        assert 'supervision_year' in config.INTEGER_FIELDS
        assert 'planned_requirement_date' in config.DATE_FIELDS
        assert 'is_blocked' in config.BOOLEAN_FIELDS

    def test_convert_cell_value_dates(self):
        """Test date conversion."""
        service = ExcelService()

        # Test datetime to date
        dt = datetime(2024, 1, 15, 10, 30, 0)
        result = service._convert_cell_value(dt, 'planned_requirement_date')
        assert result == date(2024, 1, 15)

        # Test date string parsing
        result = service._convert_cell_value('2024-01-15', 'planned_requirement_date')
        assert result == date(2024, 1, 15)

        # Test different date formats
        result = service._convert_cell_value('2024/01/15', 'planned_requirement_date')
        assert result == date(2024, 1, 15)

        # Test invalid date
        result = service._convert_cell_value('invalid-date', 'planned_requirement_date')
        assert result is None

    def test_convert_cell_value_integers(self):
        """Test integer conversion."""
        service = ExcelService()

        # Test integer
        result = service._convert_cell_value(2024, 'supervision_year')
        assert result == 2024

        # Test float to integer
        result = service._convert_cell_value(2024.0, 'supervision_year')
        assert result == 2024

        # Test string to integer
        result = service._convert_cell_value('2024', 'supervision_year')
        assert result == 2024

        # Test invalid integer
        result = service._convert_cell_value('invalid', 'supervision_year')
        assert result == 'invalid'

    def test_convert_cell_value_booleans(self):
        """Test boolean conversion."""
        service = ExcelService()

        # Test boolean values
        assert service._convert_cell_value(True, 'is_blocked') is True
        assert service._convert_cell_value(False, 'is_blocked') is False

        # Test string boolean values
        assert service._convert_cell_value('是', 'is_blocked') is True
        assert service._convert_cell_value('否', 'is_blocked') is False
        assert service._convert_cell_value('true', 'is_blocked') is True
        assert service._convert_cell_value('false', 'is_blocked') is False
        assert service._convert_cell_value('yes', 'is_blocked') is True
        assert service._convert_cell_value('no', 'is_blocked') is False

        # Test numeric boolean values
        assert service._convert_cell_value(1, 'is_blocked') is True
        assert service._convert_cell_value(0, 'is_blocked') is False

    def test_convert_cell_value_strings(self):
        """Test string conversion."""
        service = ExcelService()

        # Test string field
        result = service._convert_cell_value('Test Application', 'app_name')
        assert result == 'Test Application'

        # Test string with whitespace
        result = service._convert_cell_value('  Test App  ', 'app_name')
        assert result == 'Test App'

        # Test None value
        result = service._convert_cell_value(None, 'app_name')
        assert result is None

        # Test empty string
        result = service._convert_cell_value('', 'app_name')
        assert result is None

    @pytest.mark.asyncio
    async def test_validate_applications_data_success(self):
        """Test successful application data validation."""
        service = ExcelService()

        # Create valid test data
        test_data = {
            'l2_id': ['L2_APP_001', 'L2_APP_002'],
            'app_name': ['Test App 1', 'Test App 2'],
            'supervision_year': [2024, 2024],
            'transformation_target': ['AK', '云原生'],
            'responsible_team': ['Team A', 'Team B'],
            'overall_status': ['待启动', '研发进行中'],
            'progress_percentage': [0, 50]
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_applications_data(self.mock_db, df)

        assert len(errors) == 0

    @pytest.mark.asyncio
    async def test_validate_applications_data_missing_required_fields(self):
        """Test validation with missing required fields."""
        service = ExcelService()

        # Create data with missing required fields
        test_data = {
            'l2_id': ['L2_APP_001', ''],  # Second row missing L2 ID
            'app_name': ['Test App 1', 'Test App 2'],
            'supervision_year': [2024, None],  # Second row missing year
            'transformation_target': ['AK', '云原生'],
            'responsible_team': ['Team A', '']  # Second row missing team
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_applications_data(self.mock_db, df)

        # Should have 3 errors for row 2 (Excel row 3)
        assert len(errors) >= 3
        error_messages = [e['message'] for e in errors]
        assert any('必填字段不能为空' in msg for msg in error_messages)

    @pytest.mark.asyncio
    async def test_validate_applications_data_invalid_values(self):
        """Test validation with invalid values."""
        service = ExcelService()

        # Create data with invalid values
        test_data = {
            'l2_id': ['INVALID_ID', 'L2_APP_002'],  # Invalid L2 ID format
            'app_name': ['Test App 1', 'Test App 2'],
            'supervision_year': [1999, 2040],  # Invalid years
            'transformation_target': ['AK', 'INVALID'],  # Invalid transformation target
            'responsible_team': ['Team A', 'Team B'],
            'overall_status': ['待启动', 'INVALID_STATUS'],  # Invalid status
            'progress_percentage': [-10, 150]  # Invalid percentages
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_applications_data(self.mock_db, df)

        assert len(errors) >= 6  # Should have multiple validation errors
        error_messages = [e['message'] for e in errors]
        assert any('L2 ID必须以"L2_"开头' in msg for msg in error_messages)
        assert any('监管年必须在2020-2030之间' in msg for msg in error_messages)
        assert any('进度百分比必须在0-100之间' in msg for msg in error_messages)

    @pytest.mark.asyncio
    async def test_validate_applications_data_duplicate_l2_ids(self):
        """Test validation with duplicate L2 IDs."""
        service = ExcelService()

        # Create data with duplicate L2 IDs
        test_data = {
            'l2_id': ['L2_APP_001', 'L2_APP_001'],  # Duplicate L2 ID
            'app_name': ['Test App 1', 'Test App 2'],
            'supervision_year': [2024, 2024],
            'transformation_target': ['AK', '云原生'],
            'responsible_team': ['Team A', 'Team B']
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_applications_data(self.mock_db, df)

        # Should have errors for duplicate L2 IDs
        assert len(errors) >= 2  # One for each duplicate row
        error_messages = [e['message'] for e in errors]
        assert any('L2 ID重复' in msg for msg in error_messages)

    @pytest.mark.asyncio
    async def test_validate_subtasks_data_success(self):
        """Test successful subtask data validation."""
        service = ExcelService()

        # Mock valid L2 IDs
        mock_result = Mock()
        mock_result.all.return_value = [('L2_APP_001',), ('L2_APP_002',)]
        self.mock_db.execute.return_value = mock_result

        # Create valid test data
        test_data = {
            'application_l2_id': ['L2_APP_001', 'L2_APP_002'],
            'module_name': ['Module 1', 'Module 2'],
            'sub_target': ['AK', '云原生'],
            'task_status': ['待启动', '研发进行中'],
            'progress_percentage': [0, 75],
            'is_blocked': [False, True]
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_subtasks_data(self.mock_db, df)

        assert len(errors) == 0

    @pytest.mark.asyncio
    async def test_validate_subtasks_data_invalid_application(self):
        """Test subtask validation with invalid application L2 ID."""
        service = ExcelService()

        # Mock valid L2 IDs (empty set)
        mock_result = Mock()
        mock_result.all.return_value = []
        self.mock_db.execute.return_value = mock_result

        # Create data with invalid application L2 ID
        test_data = {
            'application_l2_id': ['L2_INVALID'],
            'module_name': ['Module 1'],
            'sub_target': ['AK']
        }

        df = pd.DataFrame(test_data)
        errors = await service._validate_subtasks_data(self.mock_db, df)

        assert len(errors) >= 1
        error_messages = [e['message'] for e in errors]
        assert any('应用L2 ID不存在' in msg for msg in error_messages)

    def test_get_column_name(self):
        """Test column name retrieval."""
        service = ExcelService()

        # Test getting column name from field name
        column = service._get_column_name('l2_id', self.config.APPLICATION_FIELDS)
        assert column == 'L2 ID'

        column = service._get_column_name('app_name', self.config.APPLICATION_FIELDS)
        assert column == '应用名称'

        # Test unknown field
        column = service._get_column_name('unknown_field', self.config.APPLICATION_FIELDS)
        assert column == 'unknown_field'

    def test_format_cell_value(self):
        """Test cell value formatting for export."""
        service = ExcelService()

        # Test date formatting
        test_date = date(2024, 1, 15)
        result = service._format_cell_value(test_date, 'planned_requirement_date')
        assert result == '2024-01-15'

        # Test boolean formatting
        result = service._format_cell_value(True, 'is_blocked')
        assert result == '是'

        result = service._format_cell_value(False, 'is_blocked')
        assert result == '否'

        # Test regular value
        result = service._format_cell_value('Test Value', 'app_name')
        assert result == 'Test Value'

        # Test None value
        result = service._format_cell_value(None, 'app_name')
        assert result == ''

    def test_create_applications_template(self):
        """Test applications template creation."""
        service = ExcelService()

        workbook = Workbook()
        service._create_applications_template(workbook, include_sample=True)

        worksheet = workbook.active
        assert worksheet.title == "应用导入模板"

        # Check headers
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = list(service.config.APPLICATION_FIELDS.keys())
        assert headers == expected_headers

        # Check sample data (should be in row 2)
        sample_row = [cell.value for cell in worksheet[2]]
        assert sample_row[0] == 'L2_APP_001'  # L2 ID
        assert sample_row[1] == '支付系统'     # App name

    def test_create_subtasks_template(self):
        """Test subtasks template creation."""
        service = ExcelService()

        workbook = Workbook()
        service._create_subtasks_template(workbook, include_sample=True)

        worksheet = workbook.active
        assert worksheet.title == "子任务导入模板"

        # Check headers
        headers = [cell.value for cell in worksheet[1]]
        expected_headers = list(service.config.SUBTASK_FIELDS.keys())
        assert headers == expected_headers

        # Check sample data
        sample_row = [cell.value for cell in worksheet[2]]
        assert sample_row[0] == 'L2_APP_001'  # Application L2 ID
        assert sample_row[1] == '用户认证模块'   # Module name

    def test_create_combined_template(self):
        """Test combined template creation."""
        service = ExcelService()

        workbook = Workbook()
        service._create_combined_template(workbook, include_sample=True)

        # Should have two sheets
        sheet_names = workbook.sheetnames
        assert '应用列表' in sheet_names
        assert '子任务列表' in sheet_names

        # Check applications sheet
        apps_sheet = workbook['应用列表']
        headers = [cell.value for cell in apps_sheet[1]]
        expected_headers = list(service.config.APPLICATION_FIELDS.keys())
        assert headers == expected_headers

        # Check subtasks sheet
        subtasks_sheet = workbook['子任务列表']
        headers = [cell.value for cell in subtasks_sheet[1]]
        expected_headers = list(service.config.SUBTASK_FIELDS.keys())
        assert headers == expected_headers

    def test_find_applications_sheet(self):
        """Test finding applications worksheet."""
        service = ExcelService()

        workbook = Workbook()
        workbook.create_sheet("应用数据")
        workbook.create_sheet("其他数据")

        sheet_name = service._find_applications_sheet(workbook)
        assert sheet_name == "应用数据"

    def test_find_subtasks_sheet(self):
        """Test finding subtasks worksheet."""
        service = ExcelService()

        workbook = Workbook()
        workbook.create_sheet("子任务列表")
        workbook.create_sheet("应用列表")

        sheet_name = service._find_subtasks_sheet(workbook)
        assert sheet_name == "子任务列表"

    def test_generate_import_template_applications(self):
        """Test generating applications import template."""
        service = ExcelService()

        template_data = service.generate_import_template("applications", include_sample_data=True)
        assert isinstance(template_data, bytes)
        assert len(template_data) > 0

        # Test that it's valid Excel data
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(template_data))
        assert workbook.active.title == "应用导入模板"

    def test_generate_import_template_subtasks(self):
        """Test generating subtasks import template."""
        service = ExcelService()

        template_data = service.generate_import_template("subtasks", include_sample_data=False)
        assert isinstance(template_data, bytes)
        assert len(template_data) > 0

        # Test that it's valid Excel data
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(template_data))
        assert workbook.active.title == "子任务导入模板"

    def test_generate_import_template_combined(self):
        """Test generating combined import template."""
        service = ExcelService()

        template_data = service.generate_import_template("combined", include_sample_data=True)
        assert isinstance(template_data, bytes)
        assert len(template_data) > 0

        # Test that it's valid Excel data
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(template_data))
        sheet_names = workbook.sheetnames
        assert "应用列表" in sheet_names
        assert "子任务列表" in sheet_names

    def test_generate_import_template_invalid_type(self):
        """Test generating template with invalid type."""
        service = ExcelService()

        with pytest.raises(ValueError, match="Unknown template type"):
            service.generate_import_template("invalid_type")

    @pytest.mark.asyncio
    async def test_import_applications_data_new_records(self):
        """Test importing new application records."""
        service = ExcelService()

        # Create test data
        test_data = {
            'l2_id': ['L2_APP_001'],
            'app_name': ['Test Application'],
            'supervision_year': [2024],
            'transformation_target': ['AK'],
            'responsible_team': ['Test Team']
        }
        df = pd.DataFrame(test_data)

        # Mock database - no existing application
        mock_result = Mock()
        mock_result.scalar_one_or_none.return_value = None
        self.mock_db.execute.return_value = mock_result

        result = await service._import_applications_data(self.mock_db, df, self.mock_user)

        assert result['imported'] == 1
        assert result['updated'] == 0
        assert result['skipped'] == 0
        self.mock_db.add.assert_called_once()
        self.mock_db.commit.assert_called_once()

    @pytest.mark.asyncio
    async def test_import_applications_data_update_existing(self):
        """Test updating existing application records."""
        service = ExcelService()

        # Create test data
        test_data = {
            'l2_id': ['L2_APP_001'],
            'app_name': ['Updated Application'],
            'supervision_year': [2024],
            'transformation_target': ['AK'],
            'responsible_team': ['Test Team']
        }
        df = pd.DataFrame(test_data)

        # Mock existing application
        mock_app = Mock(spec=Application)
        mock_app.l2_id = 'L2_APP_001'
        mock_result = Mock()
        mock_result.scalar_one_or_none.return_value = mock_app
        self.mock_db.execute.return_value = mock_result

        result = await service._import_applications_data(self.mock_db, df, self.mock_user)

        assert result['imported'] == 0
        assert result['updated'] == 1
        assert result['skipped'] == 0
        self.mock_db.commit.assert_called_once()

    @pytest.mark.asyncio
    async def test_get_applications_for_export_with_filters(self):
        """Test getting applications for export with filters."""
        service = ExcelService()

        # Mock applications
        mock_apps = [Mock(spec=Application), Mock(spec=Application)]
        mock_result = Mock()
        mock_result.scalars.return_value.all.return_value = mock_apps
        self.mock_db.execute.return_value = mock_result

        # Test with filters
        filters = {
            'supervision_year': 2024,
            'responsible_team': 'Test Team'
        }

        result = await service._get_applications_for_export(
            self.mock_db,
            application_ids=None,
            filters=filters
        )

        assert len(result) == 2
        assert result == mock_apps
        self.mock_db.execute.assert_called_once()

    @pytest.mark.asyncio
    async def test_get_subtasks_for_export_with_filters(self):
        """Test getting subtasks for export with filters."""
        service = ExcelService()

        # Mock subtasks
        mock_subtasks = [Mock(spec=SubTask), Mock(spec=SubTask)]
        mock_result = Mock()
        mock_result.scalars.return_value.all.return_value = mock_subtasks
        self.mock_db.execute.return_value = mock_result

        # Test with filters
        filters = {
            'task_status': '研发进行中',
            'is_blocked': False
        }

        result = await service._get_subtasks_for_export(
            self.mock_db,
            application_id=1,
            subtask_ids=None,
            filters=filters
        )

        assert len(result) == 2
        assert result == mock_subtasks
        self.mock_db.execute.assert_called_once()


class TestExcelValidationError:
    """Test Excel validation error class."""

    def test_excel_validation_error_with_cell_info(self):
        """Test creating validation error with cell information."""
        error = ExcelValidationError(
            message="Invalid value",
            row=5,
            column="B",
            sheet="Sheet1"
        )

        assert error.message == "Invalid value"
        assert error.row == 5
        assert error.column == "B"
        assert error.sheet == "Sheet1"
        assert "Row 5, Column B: Invalid value" in str(error)

    def test_excel_validation_error_without_cell_info(self):
        """Test creating validation error without cell information."""
        error = ExcelValidationError(message="General error")

        assert error.message == "General error"
        assert error.row is None
        assert error.column is None
        assert error.sheet is None
        assert str(error) == "General error"