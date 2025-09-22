"""
Global test configuration and fixtures for the AK Cloud Native Management System
"""

import asyncio
import pytest
import pytest_asyncio
from datetime import datetime, date, timezone
from typing import AsyncGenerator, Generator
from unittest.mock import AsyncMock, Mock, patch
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy.pool import StaticPool
from fastapi.testclient import TestClient
from httpx import AsyncClient
import tempfile
import os

from app.main import app
from app.api.deps import get_db
from app.core.config import Settings
from app.models.user import User, UserRole
from app.models.application import Application, ApplicationStatus, TransformationTarget
from app.models.subtask import SubTask, SubTaskStatus
from app.models.audit_log import AuditLog
from app.models.notification import Notification
from app.api.deps import get_current_active_user


# Test database configuration
TEST_DATABASE_URL = "sqlite+aiosqlite:///:memory:"


@pytest.fixture(scope="session")
def event_loop():
    """Create an instance of the default event loop for the test session."""
    loop = asyncio.new_event_loop()
    yield loop
    loop.close()


@pytest.fixture
async def test_engine():
    """Create test database engine."""
    engine = create_async_engine(
        TEST_DATABASE_URL,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        echo=False
    )

    # Import all models to ensure they are registered with SQLAlchemy
    from app.models import User, Application, SubTask, AuditLog, Notification
    from app.core.database import Base

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    yield engine

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    await engine.dispose()


@pytest.fixture
async def test_session(test_engine) -> AsyncGenerator[AsyncSession, None]:
    """Create test database session."""
    async_session = async_sessionmaker(
        test_engine, class_=AsyncSession, expire_on_commit=False
    )

    async with async_session() as session:
        yield session
        await session.rollback()


@pytest.fixture
def override_get_db(test_session):
    """Override database dependency for tests."""
    async def _override_get_db():
        yield test_session

    app.dependency_overrides[get_db] = _override_get_db
    yield
    app.dependency_overrides.clear()


@pytest.fixture
def test_settings():
    """Test application settings."""
    return Settings(
        APP_NAME="Test AK Cloud Native Management System",
        DEBUG=True,
        ENVIRONMENT="test",
        DATABASE_URL=TEST_DATABASE_URL,
        SECRET_KEY="test-secret-key-for-testing-only",
        JWT_SECRET_KEY="test-jwt-secret-key",
        JWT_ALGORITHM="HS256",
        JWT_EXPIRATION_HOURS=24,
        SSO_ENABLED=False
    )


@pytest.fixture
def client(override_get_db) -> Generator[TestClient, None, None]:
    """Create test client."""
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture
async def async_client(override_get_db) -> AsyncGenerator[AsyncClient, None]:
    """Create async test client."""
    async with AsyncClient(app=app, base_url="http://test") as ac:
        yield ac


# User fixtures
@pytest.fixture
def admin_user() -> User:
    """Create admin user fixture."""
    return User(
        id=1,
        sso_user_id="admin_sso_123",
        username="admin",
        full_name="Admin User",
        email="admin@example.com",
        role=UserRole.ADMIN,
        department="IT",
        is_active=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def manager_user() -> User:
    """Create manager user fixture."""
    return User(
        id=2,
        sso_user_id="manager_sso_456",
        username="manager",
        full_name="Manager User",
        email="manager@example.com",
        role=UserRole.MANAGER,
        department="Development",
        is_active=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def editor_user() -> User:
    """Create editor user fixture."""
    return User(
        id=3,
        sso_user_id="editor_sso_789",
        username="editor",
        full_name="Editor User",
        email="editor@example.com",
        role=UserRole.EDITOR,
        department="Development",
        is_active=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def viewer_user() -> User:
    """Create viewer user fixture."""
    return User(
        id=4,
        sso_user_id="viewer_sso_012",
        username="viewer",
        full_name="Viewer User",
        email="viewer@example.com",
        role=UserRole.VIEWER,
        department="Business",
        is_active=True,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def sample_application(admin_user) -> Application:
    """Create sample application fixture."""
    return Application(
        id=1,
        l2_id="L2_TEST_001",
        app_name="Test Application",
        supervision_year=2024,
        transformation_target=TransformationTarget.AK,
        responsible_team="Core Development",
        responsible_person="John Doe",
        overall_status=ApplicationStatus.NOT_STARTED,
        progress_percentage=0,
        is_ak_completed=False,
        is_cloud_native_completed=False,
        is_delayed=False,
        delay_days=0,
        planned_requirement_date=date(2024, 3, 1),
        planned_release_date=date(2024, 6, 1),
        planned_tech_online_date=date(2024, 7, 1),
        planned_biz_online_date=date(2024, 8, 1),
        notes="Test application notes",
        created_by=admin_user.id,
        updated_by=admin_user.id,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def sample_subtask(sample_application, editor_user) -> SubTask:
    """Create sample subtask fixture."""
    return SubTask(
        id=1,
        application_id=sample_application.id,
        module_name="Authentication Module",
        sub_target="SSO Integration",
        task_status=SubTaskStatus.NOT_STARTED,
        progress_percentage=0,
        planned_start_date=date(2024, 3, 1),
        planned_end_date=date(2024, 4, 1),
        is_blocked=False,
        assigned_to=editor_user.id,
        created_by=editor_user.id,
        updated_by=editor_user.id,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc)
    )


@pytest.fixture
def sample_audit_log(admin_user, sample_application) -> AuditLog:
    """Create sample audit log fixture."""
    return AuditLog(
        id=1,
        table_name="applications",
        record_id=sample_application.id,
        action="CREATE",
        old_values={},
        new_values={
            "l2_id": "L2_TEST_001",
            "app_name": "Test Application",
            "supervision_year": 2024
        },
        user_id=admin_user.id,
        timestamp=datetime.now(timezone.utc),
        request_id="test-request-123"
    )


@pytest.fixture
def sample_notification(manager_user) -> Notification:
    """Create sample notification fixture."""
    return Notification(
        id=1,
        user_id=manager_user.id,
        title="Test Notification",
        message="This is a test notification",
        type="delay_warning",
        is_read=False,
        created_at=datetime.now(timezone.utc)
    )


# Authentication fixtures
@pytest.fixture
def mock_admin_auth(admin_user):
    """Mock admin user authentication."""
    with patch('app.api.deps.get_current_active_user', return_value=admin_user):
        yield admin_user


@pytest.fixture
def mock_manager_auth(manager_user):
    """Mock manager user authentication."""
    with patch('app.api.deps.get_current_active_user', return_value=manager_user):
        yield manager_user


@pytest.fixture
def mock_editor_auth(editor_user):
    """Mock editor user authentication."""
    with patch('app.api.deps.get_current_active_user', return_value=editor_user):
        yield editor_user


@pytest.fixture
def mock_viewer_auth(viewer_user):
    """Mock viewer user authentication."""
    with patch('app.api.deps.get_current_active_user', return_value=viewer_user):
        yield viewer_user


# Database fixtures with data
@pytest.fixture
async def db_with_users(test_session, admin_user, manager_user, editor_user, viewer_user):
    """Populate database with sample users."""
    users = [admin_user, manager_user, editor_user, viewer_user]
    for user in users:
        test_session.add(user)

    await test_session.commit()

    for user in users:
        await test_session.refresh(user)

    return users


@pytest.fixture
async def db_with_application(test_session, db_with_users, sample_application):
    """Populate database with sample application."""
    test_session.add(sample_application)
    await test_session.commit()
    await test_session.refresh(sample_application)
    return sample_application


@pytest.fixture
async def db_with_subtask(test_session, db_with_application, sample_subtask):
    """Populate database with sample subtask."""
    test_session.add(sample_subtask)
    await test_session.commit()
    await test_session.refresh(sample_subtask)
    return sample_subtask


@pytest.fixture
async def db_with_full_data(test_session, db_with_subtask, sample_audit_log, sample_notification):
    """Populate database with complete sample data."""
    test_session.add(sample_audit_log)
    test_session.add(sample_notification)
    await test_session.commit()
    await test_session.refresh(sample_audit_log)
    await test_session.refresh(sample_notification)
    return {
        "users": await test_session.get(User, 1),
        "application": await test_session.get(Application, 1),
        "subtask": await test_session.get(SubTask, 1),
        "audit_log": sample_audit_log,
        "notification": sample_notification
    }


# Mock services fixtures
@pytest.fixture
def mock_auth_service():
    """Mock authentication service."""
    service = Mock()
    service.verify_sso_token = AsyncMock(return_value={"user_id": "test_user_123", "email": "test@example.com"})
    service.create_access_token = Mock(return_value="mock_jwt_token")
    service.verify_token = Mock(return_value={"sub": "test_user_123"})
    return service


@pytest.fixture
def mock_calculation_engine():
    """Mock calculation engine."""
    engine = Mock()
    engine.calculate_application_progress = AsyncMock(return_value={
        "overall_status": ApplicationStatus.DEV_IN_PROGRESS,
        "progress_percentage": 45,
        "is_delayed": False,
        "delay_days": 0
    })
    engine.bulk_calculate = AsyncMock(return_value=[])
    return engine


@pytest.fixture
def mock_excel_service():
    """Mock Excel service."""
    service = Mock()
    service.import_applications = AsyncMock(return_value={"imported": 5, "errors": []})
    service.export_applications = AsyncMock(return_value=b"mock_excel_content")
    service.generate_template = Mock(return_value=b"mock_template_content")
    return service


@pytest.fixture
def mock_notification_service():
    """Mock notification service."""
    service = Mock()
    service.send_notification = AsyncMock(return_value=True)
    service.send_bulk_notifications = AsyncMock(return_value={"sent": 5, "failed": 0})
    service.create_delay_warning = AsyncMock(return_value=True)
    return service


@pytest.fixture
def mock_report_service():
    """Mock report service."""
    service = Mock()
    service.generate_progress_report = AsyncMock(return_value={
        "total_applications": 10,
        "completed": 3,
        "in_progress": 5,
        "not_started": 2
    })
    service.export_report = AsyncMock(return_value=b"mock_report_content")
    return service


# File fixtures
@pytest.fixture
def temp_excel_file():
    """Create temporary Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        # Create minimal Excel content for testing
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'L2 ID'
        ws['B1'] = 'Application Name'
        ws['A2'] = 'L2_TEST_001'
        ws['B2'] = 'Test Application'
        wb.save(tmp.name)

        yield tmp.name

        try:
            os.unlink(tmp.name)
        except OSError:
            pass


# Error simulation fixtures
@pytest.fixture
def mock_db_error():
    """Mock database error for testing error handling."""
    def _mock_error(*args, **kwargs):
        from sqlalchemy.exc import SQLAlchemyError
        raise SQLAlchemyError("Mock database error")

    return _mock_error


@pytest.fixture
def mock_timeout_error():
    """Mock timeout error for testing timeout handling."""
    def _mock_error(*args, **kwargs):
        import asyncio
        raise asyncio.TimeoutError("Mock timeout error")

    return _mock_error


# Performance testing fixtures
@pytest.fixture
def performance_monitor():
    """Monitor performance metrics during tests."""
    import time
    import psutil

    class PerformanceMonitor:
        def __init__(self):
            self.start_time = None
            self.start_memory = None

        def start(self):
            self.start_time = time.time()
            self.start_memory = psutil.Process().memory_info().rss

        def stop(self):
            elapsed_time = time.time() - self.start_time
            current_memory = psutil.Process().memory_info().rss
            memory_delta = current_memory - self.start_memory

            return {
                "elapsed_time": elapsed_time,
                "memory_delta": memory_delta,
                "current_memory": current_memory
            }

    return PerformanceMonitor()


# Parametrized fixtures for comprehensive testing
@pytest.fixture(params=[
    ApplicationStatus.NOT_STARTED,
    ApplicationStatus.DEV_IN_PROGRESS,
    ApplicationStatus.BIZ_ONLINE,
    ApplicationStatus.COMPLETED
])
def application_status(request):
    """Parametrized application status for comprehensive testing."""
    return request.param


@pytest.fixture(params=[
    SubTaskStatus.NOT_STARTED,
    SubTaskStatus.DEV_IN_PROGRESS,
    SubTaskStatus.TECH_ONLINE,
    SubTaskStatus.BIZ_ONLINE,
    SubTaskStatus.COMPLETED
])
def subtask_status(request):
    """Parametrized subtask status for comprehensive testing."""
    return request.param


@pytest.fixture(params=[
    TransformationTarget.AK,
    TransformationTarget.CLOUD_NATIVE
])
def transformation_target(request):
    """Parametrized transformation target for comprehensive testing."""
    return request.param


@pytest.fixture(params=[
    UserRole.ADMIN,
    UserRole.MANAGER,
    UserRole.EDITOR,
    UserRole.VIEWER
])
def user_role(request):
    """Parametrized user role for comprehensive testing."""
    return request.param


# Cleanup fixture
@pytest.fixture(autouse=True)
def cleanup_after_test():
    """Cleanup after each test."""
    yield
    # Cleanup any test artifacts, temporary files, etc.
    pass