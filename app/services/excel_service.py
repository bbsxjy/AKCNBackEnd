"""
Excel Import/Export Service

Provides comprehensive Excel processing capabilities for applications and subtasks
with template-based imports, bulk data validation, and error reporting.
"""

import io
import json
from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime, date, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from app.models.application import Application, ApplicationStatus, TransformationTarget
from app.models.subtask import SubTask, SubTaskStatus
from app.models.user import User
from app.core.exceptions import ValidationError, BusinessLogicError


class ExcelValidationError(Exception):
    """Excel validation error with cell-level details."""

    def __init__(self, message: str, row: int = None, column: str = None, sheet: str = None):
        self.message = message
        self.row = row
        self.column = column
        self.sheet = sheet
        super().__init__(f"Row {row}, Column {column}: {message}" if row and column else message)


class ExcelMappingConfig:
    """Configuration for Excel field mappings."""

    # Application field mappings (支持前端发送的英文字段名)
    APPLICATION_FIELDS = {
        # 前端发送的英文字段名 - 保留原有的英文字段映射
        'application_id': 'l2_id',
        'l2_id': 'l2_id',
        'application_name': 'app_name',
        'app_name': 'app_name',

        # 用户指定的精确中文列名映射
        'L2ID': 'l2_id',
        'L2应用': 'app_name',
        'L2应用名': 'app_name',
        '监管验收年份': 'ak_supervision_acceptance_year',
        '改造目标': 'overall_transformation_target',
        '是否已完成AK': 'is_ak_completed',
        '是否已完成云原生': 'is_cloud_native_completed',
        '当前改造阶段': 'current_transformation_phase',
        '改造状态': 'current_status',
        '【计划】\n需求完成时间': 'planned_requirement_date',
        '【计划】需求完成时间': 'planned_requirement_date',  # 无换行版本
        '【计划】\n发版时间': 'planned_release_date',
        '【计划】发版时间': 'planned_release_date',  # 无换行版本
        '【计划】\n技术上线时间': 'planned_tech_online_date',
        '【计划】技术上线时间': 'planned_tech_online_date',  # 无换行版本
        '【计划】\n业务上线时间': 'planned_biz_online_date',
        '【计划】业务上线时间': 'planned_biz_online_date',  # 无换行版本
        '【实际】\n需求到达时间': 'actual_requirement_date',
        '【实际】需求到达时间': 'actual_requirement_date',  # 无换行版本
        '【实际】\n发版时间': 'actual_release_date',
        '【实际】发版时间': 'actual_release_date',  # 无换行版本
        '【实际】\n技术上线时间': 'actual_tech_online_date',
        '【实际】技术上线时间': 'actual_tech_online_date',  # 无换行版本
        '【实际】\n业务上线时间': 'actual_biz_online_date',
        '【实际】业务上线时间': 'actual_biz_online_date',  # 无换行版本
        '备注': 'notes',
        '档位': 'app_tier',
        '所属L1': 'belonging_l1_name',
        '所属项目': 'belonging_projects',
        '开发模式': 'dev_mode',
        '运维模式': 'ops_mode',
        '开发负责人': 'dev_owner',
        '开发团队': 'dev_team',
        '运维负责人': 'ops_owner',
        '运维团队': 'ops_team',
        '所属指标': 'belonging_kpi',
        '验收状态': 'acceptance_status',

        # Old field names mapping to new database columns (保留英文字段支持)
        'supervision_year': 'ak_supervision_acceptance_year',
        'ak_supervision_acceptance_year': 'ak_supervision_acceptance_year',
        'transformation_target': 'overall_transformation_target',
        'overall_transformation_target': 'overall_transformation_target',
        'current_stage': 'current_transformation_phase',
        'current_transformation_phase': 'current_transformation_phase',
        'status': 'current_status',
        'overall_status': 'current_status',
        'current_status': 'current_status',

        # Team and ownership fields (英文支持)
        'responsible_team': 'dev_team',
        'dev_team': 'dev_team',
        'ops_team': 'ops_team',
        'responsible_person': 'dev_owner',
        'dev_owner': 'dev_owner',
        'ops_owner': 'ops_owner',

        # New fields (英文支持)
        'app_tier': 'app_tier',
        'belonging_l1_name': 'belonging_l1_name',
        'belonging_projects': 'belonging_projects',
        'is_ak_completed': 'is_ak_completed',
        'is_cloud_native_completed': 'is_cloud_native_completed',
        'is_domain_transformation_completed': 'is_domain_transformation_completed',
        'is_dbpm_transformation_completed': 'is_dbpm_transformation_completed',
        'dev_mode': 'dev_mode',
        'ops_mode': 'ops_mode',
        'belonging_kpi': 'belonging_kpi',
        'acceptance_status': 'acceptance_status',
        'is_delayed': 'is_delayed',
        'delay_days': 'delay_days',
        'planned_requirement_date': 'planned_requirement_date',
        'planned_release_date': 'planned_release_date',
        'planned_tech_online_date': 'planned_tech_online_date',
        'planned_biz_online_date': 'planned_biz_online_date',
        'actual_requirement_date': 'actual_requirement_date',
        'actual_release_date': 'actual_release_date',
        'actual_tech_online_date': 'actual_tech_online_date',
        'actual_biz_online_date': 'actual_biz_online_date',
        'notes': 'notes',

        # 保留其他可能的中文变体（作为备用）
        'L2 ID': 'l2_id',
        'L2-ID': 'l2_id',
        'L2_ID': 'l2_id',
        'L2编号': 'l2_id',
        '应用编号': 'l2_id',
        '系统编号': 'l2_id',
        '序号': 'l2_id',

        '应用名称': 'app_name',
        '应用名': 'app_name',
        '系统名称': 'app_name',
        '系统名': 'app_name',
        'L2应用名': 'app_name',
        'L2应用名称': 'app_name',
        '名称': 'app_name',

        '监管年': 'ak_supervision_acceptance_year',
        '监管年度': 'ak_supervision_acceptance_year',
        '监管年份': 'ak_supervision_acceptance_year',
        '年度': 'ak_supervision_acceptance_year',
        '指标年度': 'ak_supervision_acceptance_year',
        '指标标签': 'ak_supervision_acceptance_year',
        'AK监管验收年': 'ak_supervision_acceptance_year',
        '验收年度': 'ak_supervision_acceptance_year',
        '验收年份': 'ak_supervision_acceptance_year',

        '转型目标': 'overall_transformation_target',
        '目标': 'overall_transformation_target',
        '改造类型': 'overall_transformation_target',
        'AK/云原生': 'overall_transformation_target',
        'AK/Cloud': 'overall_transformation_target',
        '改造方向': 'overall_transformation_target',
        '整体转型目标': 'overall_transformation_target',

        '当前阶段': 'current_transformation_phase',
        '当前状态': 'current_transformation_phase',
        '阶段': 'current_transformation_phase',
        '进展阶段': 'current_transformation_phase',
        '开发阶段': 'current_transformation_phase',
        '当前转型阶段': 'current_transformation_phase',

        '整体状态': 'current_status',
        '状态': 'current_status',
        '总体状态': 'current_status',
        '完成状态': 'current_status',
        '整体进展': 'current_status',
        '当前状态': 'current_status',

        '负责团队': 'dev_team',
        '团队': 'dev_team',
        '改造团队': 'dev_team',
        '负责部门': 'dev_team',
        '责任团队': 'dev_team',
        '所属团队': 'dev_team',

        '负责人': 'dev_owner',
        '责任人': 'dev_owner',
        '项目负责人': 'dev_owner',
        '团队负责人': 'dev_owner',
        '联系人': 'dev_owner',

        # Additional mappings
        '是否已完成域名化改造': 'is_domain_transformation_completed',
        '是否已完成DBPM改造': 'is_dbpm_transformation_completed',
        '研发模式': 'dev_mode',
        '是否完成验收': 'acceptance_status',

        # 其他组织字段的中文映射
        '应用层级': 'app_tier',
        '所属L1名称': 'belonging_l1_name',
        'AK完成': 'is_ak_completed',
        '云原生完成': 'is_cloud_native_completed',
        '领域转型完成': 'is_domain_transformation_completed',
        'DBPM转型完成': 'is_dbpm_transformation_completed',
        '所属KPI': 'belonging_kpi',
        '是否延期': 'is_delayed',
        '延期天数': 'delay_days',

        '计划需求日期': 'planned_requirement_date',
        '计划需求': 'planned_requirement_date',
        '需求计划': 'planned_requirement_date',

        '计划发布日期': 'planned_release_date',
        '计划发布': 'planned_release_date',
        '发布计划': 'planned_release_date',

        '计划技术上线日期': 'planned_tech_online_date',
        '计划技术上线': 'planned_tech_online_date',
        '技术上线计划': 'planned_tech_online_date',

        '计划业务上线日期': 'planned_biz_online_date',
        '计划业务上线': 'planned_biz_online_date',
        '业务上线计划': 'planned_biz_online_date',
        '计划上线': 'planned_biz_online_date',
        '计划完成日期': 'planned_biz_online_date',

        '实际需求日期': 'actual_requirement_date',
        '实际需求': 'actual_requirement_date',
        '需求实际': 'actual_requirement_date',

        '实际发布日期': 'actual_release_date',
        '实际发布': 'actual_release_date',
        '发布实际': 'actual_release_date',

        '实际技术上线日期': 'actual_tech_online_date',
        '实际技术上线': 'actual_tech_online_date',
        '技术上线实际': 'actual_tech_online_date',

        '实际业务上线日期': 'actual_biz_online_date',
        '实际业务上线': 'actual_biz_online_date',
        '业务上线实际': 'actual_biz_online_date',
        '实际上线': 'actual_biz_online_date',
        '实际完成日期': 'actual_biz_online_date',

        '备注': 'notes',
        '说明': 'notes',
        '描述': 'notes',
        '注释': 'notes',
        '备注说明': 'notes',
        '其他': 'notes'
    }

    # SubTask field mappings (支持前端发送的英文字段名)
    SUBTASK_FIELDS = {
        # 用户指定的精确中文列名映射 (子追踪表)
        'L2ID': 'l2_id',
        'L2应用': 'app_name',
        'L2应用名': 'app_name',
        '子目标': 'sub_target',
        '版本名': 'version_name',
        '改造状态': 'task_status',
        '【计划】\n需求完成时间': 'planned_requirement_date',
        '【计划】需求完成时间': 'planned_requirement_date',  # 无换行版本
        '【计划】\n发版时间': 'planned_release_date',
        '【计划】发版时间': 'planned_release_date',  # 无换行版本
        '【计划】\n技术上线时间': 'planned_tech_online_date',
        '【计划】技术上线时间': 'planned_tech_online_date',  # 无换行版本
        '【计划】\n业务上线时间': 'planned_biz_online_date',
        '【计划】业务上线时间': 'planned_biz_online_date',  # 无换行版本
        '【实际】\n需求到达时间': 'actual_requirement_date',
        '【实际】需求到达时间': 'actual_requirement_date',  # 无换行版本
        '【实际】\n发版时间': 'actual_release_date',
        '【实际】发版时间': 'actual_release_date',  # 无换行版本
        '【实际】\n技术上线时间': 'actual_tech_online_date',
        '【实际】技术上线时间': 'actual_tech_online_date',  # 无换行版本
        '【实际】\n业务上线时间': 'actual_biz_online_date',
        '【实际】业务上线时间': 'actual_biz_online_date',  # 无换行版本
        '备注': 'notes',
        '资源是否申请': 'resource_applied',
        '运营需求提交': 'ops_requirement_submitted',
        '运营测试': 'ops_testing_status',
        '上线检查': 'launch_check_status',

        # 前端发送的英文字段名 (保留英文支持)
        'application_l2_id': 'l2_id',
        'app_l2_id': 'l2_id',
        'l2_id': 'l2_id',
        'app_name': 'app_name',
        'sub_target': 'sub_target',
        'target': 'sub_target',
        'transformation_target': 'sub_target',
        'version_name': 'version_name',
        'version': 'version_name',
        'task_status': 'task_status',
        'status': 'task_status',
        'progress_percentage': 'progress_percentage',
        'progress': 'progress_percentage',
        'is_blocked': 'is_blocked',
        'blocked': 'is_blocked',
        'block_reason': 'block_reason',
        'planned_requirement_date': 'planned_requirement_date',
        'planned_release_date': 'planned_release_date',
        'planned_tech_online_date': 'planned_tech_online_date',
        'planned_biz_online_date': 'planned_biz_online_date',
        'actual_requirement_date': 'actual_requirement_date',
        'actual_release_date': 'actual_release_date',
        'actual_tech_online_date': 'actual_tech_online_date',
        'actual_biz_online_date': 'actual_biz_online_date',
        'notes': 'notes',
        'remarks': 'notes',
        'description': 'notes',
        'technical_notes': 'notes',

        # New fields (英文支持)
        'resource_applied': 'resource_applied',
        'ops_requirement_submitted': 'ops_requirement_submitted',
        'ops_testing_status': 'ops_testing_status',
        'launch_check_status': 'launch_check_status',

        # 保留其他可能的中文变体（作为备用）
        'L2 ID': 'l2_id',
        '应用L2 ID': 'l2_id',
        '应用名称': 'app_name',
        '应用名': 'app_name',
        'L2应用名': 'app_name',
        '版本名称': 'version_name',
        '任务状态': 'task_status',
        '子任务完成': 'task_status',
        '进度百分比': 'progress_percentage',
        '是否阻塞': 'is_blocked',
        '阻塞原因': 'block_reason',
        '计划需求日期': 'planned_requirement_date',
        '计划发布日期': 'planned_release_date',
        '计划技术上线日期': 'planned_tech_online_date',
        '计划业务上线日期': 'planned_biz_online_date',
        '实际需求日期': 'actual_requirement_date',
        '实际发布日期': 'actual_release_date',
        '实际技术上线日期': 'actual_tech_online_date',
        '实际业务上线日期': 'actual_biz_online_date',
        '子表备注': 'notes',
        '主表同步备注': 'notes',

        # Other field variants
        '资源已申请': 'resource_applied',
        '运维需求提交时间': 'ops_requirement_submitted',
        '运维测试状态': 'ops_testing_status',
        '运营测试状态': 'ops_testing_status',
        '上线检查状态': 'launch_check_status'
    }

    # Required fields (调整为更宽松的验证，适配前端数据)
    APPLICATION_REQUIRED = ['l2_id']  # 只要求L2 ID为必填，其他字段可以为空
    SUBTASK_REQUIRED = ['l2_id']  # 只要求L2 ID为必填，其他字段可以为空并设置默认值

    # Data type mappings
    DATE_FIELDS = [
        'planned_requirement_date', 'planned_release_date', 'planned_tech_online_date', 'planned_biz_online_date',
        'actual_requirement_date', 'actual_release_date', 'actual_tech_online_date', 'actual_biz_online_date'
    ]

    DATETIME_FIELDS = ['ops_requirement_submitted']  # Timestamp fields

    INTEGER_FIELDS = [
        'ak_supervision_acceptance_year', 'app_tier', 'progress_percentage', 'delay_days'
    ]

    BOOLEAN_FIELDS = [
        'is_blocked', 'is_ak_completed', 'is_cloud_native_completed',
        'is_domain_transformation_completed', 'is_dbpm_transformation_completed',
        'is_delayed', 'resource_applied'
    ]

    STRING_FIELDS = [
        'l2_id', 'app_name', 'dev_team', 'ops_team', 'dev_owner', 'ops_owner',
        'belonging_l1_name', 'belonging_projects', 'dev_mode', 'ops_mode',
        'belonging_kpi', 'acceptance_status', 'notes', 'current_status',
        'overall_transformation_target', 'current_transformation_phase',
        'sub_target', 'version_name', 'task_status', 'block_reason'
    ]


class ExcelService:
    """Service for Excel import/export operations."""

    def __init__(self):
        self.config = ExcelMappingConfig()

    # Import Operations

    async def import_applications_from_excel(
        self,
        db: AsyncSession,
        file_content: bytes,
        user: User,
        validate_only: bool = False
    ) -> Dict[str, Any]:
        """Import applications from Excel file."""

        try:
            print(f"[INFO] Starting import_applications_from_excel")
            
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            print(f"[INFO] Loaded workbook with sheets: {workbook.sheetnames}")

            # Process the first worksheet or find Applications sheet
            sheet_name = self._find_applications_sheet(workbook)
            print(f"[INFO] Using applications sheet: {sheet_name}")
            worksheet = workbook[sheet_name]

            # Convert to DataFrame
            df = self._worksheet_to_dataframe(worksheet, self.config.APPLICATION_FIELDS)
            print(f"[INFO] Applications DataFrame shape: {df.shape}")
            print(f"[INFO] Applications DataFrame columns: {list(df.columns)}")

            # Validate data
            validation_errors = await self._validate_applications_data(db, df)
            print(f"[INFO] Validation errors count: {len(validation_errors)}")
            if validation_errors:
                print(f"[WARNING] First few validation errors: {validation_errors[:3]}")

            if validation_errors and not validate_only:
                return {
                    'success': False,
                    'errors': validation_errors,
                    'total_rows': len(df),
                    'processed_rows': 0
                }

            if validate_only:
                return {
                    'success': len(validation_errors) == 0,
                    'errors': validation_errors,
                    'total_rows': len(df),
                    'preview_data': df.head(5).to_dict('records') if not validation_errors else []
                }

            # Additional debugging before import
            print(f"[INFO] About to import {len(df)} application rows")
            if 'app_name' in df.columns and len(df) > 0:
                print(f"[DEBUG] First 3 app_name values: {list(df['app_name'].head(3))}")

            # Import data
            print(f"[INFO] Starting data import...")
            results = await self._import_applications_data(db, df, user)
            print(f"[INFO] Import completed successfully")
            print(f"[INFO] Results: imported={results['imported']}, updated={results['updated']}, skipped={results['skipped']}")

            return {
                'success': True,
                'total_rows': len(df),
                'processed_rows': results['imported'],
                'updated_rows': results['updated'],
                'skipped_rows': results['skipped'],
                'errors': validation_errors
            }

        except Exception as e:
            print(f"[ERROR] Exception in import_applications_from_excel: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e),
                'total_rows': 0,
                'processed_rows': 0
            }

    async def import_subtasks_from_excel(
        self,
        db: AsyncSession,
        file_content: bytes,
        user: User,
        validate_only: bool = False
    ) -> Dict[str, Any]:
        """Import subtasks from Excel file with support for two-sheet import."""

        try:
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            print(f"DEBUG: Loaded workbook with sheets: {workbook.sheetnames}")

            # Check if this is a two-sheet import (applications + subtasks)
            has_applications_sheet = any(keyword in sheet_name.lower() for sheet_name in workbook.sheetnames
                                       for keyword in ['总追踪表', '应用', 'application', 'app'])
            has_subtasks_sheet = any(keyword in sheet_name.lower() for sheet_name in workbook.sheetnames
                                   for keyword in ['子追踪表', '子任务', 'subtask', 'task'])

            print(f"DEBUG: has_applications_sheet: {has_applications_sheet}")
            print(f"DEBUG: has_subtasks_sheet: {has_subtasks_sheet}")

            total_app_rows = 0
            total_subtask_rows = 0
            all_validation_errors = []
            app_results = {'imported': 0, 'updated': 0, 'skipped': 0}
            subtask_results = {'imported': 0, 'updated': 0, 'skipped': 0}

            # Initialize DataFrames
            app_df = pd.DataFrame()

            # Import applications first if both sheets exist
            if has_applications_sheet:
                app_sheet_name = self._find_applications_sheet(workbook)
                app_worksheet = workbook[app_sheet_name]
                app_df = self._worksheet_to_dataframe(app_worksheet, self.config.APPLICATION_FIELDS)

                if len(app_df) > 0:
                    total_app_rows = len(app_df)
                    app_validation_errors = await self._validate_applications_data(db, app_df)
                    all_validation_errors.extend([{**error, 'sheet': '总追踪表'} for error in app_validation_errors])

                    if not validate_only and len(app_validation_errors) == 0:
                        print(f"[INFO] Starting import of {total_app_rows} applications...")
                        try:
                            app_results = await self._import_applications_data(db, app_df, user)
                            print(f"[INFO] Applications import completed: {app_results}")
                        except Exception as e:
                            print(f"[ERROR] Failed to import applications: {e}")
                            import traceback
                            traceback.print_exc()
                            raise
                else:
                    print(f"DEBUG: No application data found in worksheet {app_sheet_name}")

            # Import subtasks
            if has_subtasks_sheet:
                subtask_sheet_name = self._find_subtasks_sheet(workbook)
                print(f"DEBUG: Using subtasks sheet: {subtask_sheet_name}")
                subtask_worksheet = workbook[subtask_sheet_name]
            else:
                # Fallback to first sheet if no specific subtask sheet found
                subtask_worksheet = workbook.active
                print(f"DEBUG: Using default active sheet: {subtask_worksheet.title}")

            print(f"DEBUG: Processing subtask worksheet...")
            subtask_df = self._worksheet_to_dataframe(subtask_worksheet, self.config.SUBTASK_FIELDS)
            print(f"DEBUG: SubTask DataFrame shape: {subtask_df.shape}")
            print(f"DEBUG: SubTask DataFrame columns: {list(subtask_df.columns)}")

            if len(subtask_df) > 0:
                total_subtask_rows = len(subtask_df)
                print(f"DEBUG: Found {total_subtask_rows} subtask rows")
                subtask_validation_errors = await self._validate_subtasks_data(db, subtask_df)
                all_validation_errors.extend([{**error, 'sheet': '子追踪表'} for error in subtask_validation_errors])

                if not validate_only and len(subtask_validation_errors) == 0:
                    subtask_results = await self._import_subtasks_data(db, subtask_df, user)
            else:
                print(f"DEBUG: No subtask data found in worksheet")
                # If no data was found and no application data either, this is an error
                if total_app_rows == 0:
                    return {
                        'success': False,
                        'error': 'No recognizable data found in Excel file. Please check column headers match expected format.',
                        'total_rows': 0,
                        'processed_rows': 0
                    }

            total_rows = total_app_rows + total_subtask_rows

            if all_validation_errors and not validate_only:
                return {
                    'success': False,
                    'errors': all_validation_errors,
                    'total_rows': total_rows,
                    'processed_rows': 0
                }

            if validate_only:
                preview_data = []
                if total_app_rows > 0:
                    preview_data.extend([{**row, 'sheet': '总追踪表'} for row in app_df.head(3).to_dict('records')])
                if total_subtask_rows > 0:
                    preview_data.extend([{**row, 'sheet': '子追踪表'} for row in subtask_df.head(3).to_dict('records')])

                return {
                    'success': len(all_validation_errors) == 0,
                    'errors': all_validation_errors,
                    'total_rows': total_rows,
                    'preview_data': preview_data if len(all_validation_errors) == 0 else []
                }

            # Combine results for response
            print(f"[INFO] Combining results...")
            total_imported = app_results['imported'] + subtask_results['imported']
            total_updated = app_results['updated'] + subtask_results['updated']
            total_skipped = app_results['skipped'] + subtask_results['skipped']
            print(f"[INFO] Total results: imported={total_imported}, updated={total_updated}, skipped={total_skipped}")

            return {
                'success': True,
                'total_rows': total_rows,
                'processed_rows': total_imported,
                'updated_rows': total_updated,
                'skipped_rows': total_skipped,
                'applications': {
                    'total_rows': total_app_rows,
                    'imported': app_results['imported'],
                    'updated': app_results['updated'],
                    'skipped': app_results['skipped']
                },
                'subtasks': {
                    'total_rows': total_subtask_rows,
                    'imported': subtask_results['imported'],
                    'updated': subtask_results['updated'],
                    'skipped': subtask_results['skipped']
                },
                'errors': all_validation_errors
            }

        except Exception as e:
            print(f"[ERROR] Exception in import_subtasks_from_excel: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e),
                'total_rows': 0,
                'processed_rows': 0
            }

    # Export Operations

    async def export_applications_to_excel(
        self,
        db: AsyncSession,
        application_ids: Optional[List[int]] = None,
        filters: Optional[Dict[str, Any]] = None,
        template_style: str = "standard"
    ) -> bytes:
        """Export applications to Excel file."""

        # Get applications data
        applications = await self._get_applications_for_export(db, application_ids, filters)

        # Create workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "应用列表"

        # Setup headers
        headers = list(self.config.APPLICATION_FIELDS.keys())
        self._write_headers(worksheet, headers, template_style)

        # Write data
        for row_num, app in enumerate(applications, start=2):
            for col_num, (header, field) in enumerate(self.config.APPLICATION_FIELDS.items(), start=1):
                value = self._format_cell_value(getattr(app, field, ''), field)
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Apply styling
        self._apply_worksheet_styling(worksheet, len(applications) + 1, len(headers), template_style)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    async def export_subtasks_to_excel(
        self,
        db: AsyncSession,
        application_id: Optional[int] = None,
        subtask_ids: Optional[List[int]] = None,
        filters: Optional[Dict[str, Any]] = None,
        template_style: str = "standard"
    ) -> bytes:
        """Export subtasks to Excel file."""

        # Get subtasks data
        subtasks = await self._get_subtasks_for_export(db, application_id, subtask_ids, filters)

        # Create workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "子任务列表"

        # Setup headers
        headers = list(self.config.SUBTASK_FIELDS.keys())
        self._write_headers(worksheet, headers, template_style)

        # Write data
        for row_num, subtask in enumerate(subtasks, start=2):
            for col_num, (header, field) in enumerate(self.config.SUBTASK_FIELDS.items(), start=1):
                if field == 'l2_id':  
                    # 获取关联的Application的l2_id
                    value = subtask.application.l2_id if subtask.application else ''
                elif field == 'app_name':
                    # 从SubTask直接获取app_name字段
                    value = subtask.app_name if hasattr(subtask, 'app_name') else ''
                else:
                    value = self._format_cell_value(getattr(subtask, field, ''), field)
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Apply styling
        self._apply_worksheet_styling(worksheet, len(subtasks) + 1, len(headers), template_style)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_import_template(
        self,
        template_type: str,
        include_sample_data: bool = False
    ) -> bytes:
        """Generate Excel import template."""

        workbook = Workbook()

        if template_type == "applications":
            self._create_applications_template(workbook, include_sample_data)
        elif template_type == "subtasks":
            self._create_subtasks_template(workbook, include_sample_data)
        elif template_type == "combined":
            self._create_combined_template(workbook, include_sample_data)
        else:
            raise ValueError(f"Unknown template type: {template_type}")

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    # Helper Methods

    def _find_applications_sheet(self, workbook) -> str:
        """Find applications worksheet."""
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['总追踪表', '应用', 'application', 'app']):
                print(f"DEBUG: Found applications sheet: {sheet_name}")

                # Preview first few rows to help identify structure
                worksheet = workbook[sheet_name]
                print(f"DEBUG: Sheet preview (first 10 rows, first 5 columns):")
                for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, max_col=5), start=1):
                    row_preview = [str(cell.value)[:30] if cell.value else 'None' for cell in row]
                    print(f"  Row {row_num}: {row_preview}")

                return sheet_name
        print(f"DEBUG: No applications sheet found, using first sheet: {workbook.sheetnames[0]}")
        return workbook.sheetnames[0]  # Default to first sheet

    def _find_subtasks_sheet(self, workbook) -> str:
        """Find subtasks worksheet."""
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['子追踪表', '子任务', 'subtask', 'task']):
                print(f"DEBUG: Found subtasks sheet: {sheet_name}")
                return sheet_name
        print(f"DEBUG: No subtasks sheet found, using first sheet: {workbook.sheetnames[0]}")
        return workbook.sheetnames[0]  # Default to first sheet

    def _worksheet_to_dataframe(self, worksheet, field_mapping: Dict[str, str]) -> pd.DataFrame:
        """Convert worksheet to DataFrame with field mapping."""

        # Get all data
        data = []
        headers = []

        # Keywords that indicate actual column headers (not instruction/statistics text)
        header_keywords = [
            'L2', 'ID', '应用', '名称', '模块', '状态', '进度', '负责', '日期', '备注',
            'application', 'module', 'status', 'progress', 'date', 'team', 'person',
            '版本', '目标', '阶段', '团队', '百分比', 'name', 'target', '序号',
            '系统', '监管', '改造', '开发', '上线', '发布', '需求', '档位', '所属',
            '模式', '指标', '验收', '【计划】', '【实际】', 'L2ID', 'L2应用', "L2应用名",
            'l2_id', 'app_name', 'sub_target', 'version_name', 'task_status'  # Add English field names
        ]

        # Keywords that indicate statistics/summary rows (should be skipped)
        skip_keywords = [
            '本月计划', '分项统计', '已完成', '工作进度', '统计', '汇总', '合计',
            '总计', '小计', '说明', '使用说明', '指标标签解释', '表格使用说明'
        ]

        # Find header row by checking for actual column names, not instruction/statistics text
        header_row = 1
        best_match = {'row': 1, 'score': 0, 'headers': []}

        for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=30), start=1):
            row_values = [str(cell.value).strip() if cell.value else '' for cell in row]
            non_empty_values = [v for v in row_values if v]

            # Skip rows that look like instructions (very long text in first cell)
            if non_empty_values and len(non_empty_values[0]) > 100:
                print(f"DEBUG: Row {row_num} looks like instructions, skipping...")
                continue

            # Skip rows that contain statistics/summary keywords
            if non_empty_values:
                row_text = ' '.join(non_empty_values[:5]).lower()  # Check first 5 cells
                if any(skip_word in row_text for skip_word in skip_keywords):
                    print(f"DEBUG: Row {row_num} looks like statistics/summary, skipping...")
                    continue

            # Check if this row looks like data (has many numbers/dates)
            numeric_count = 0
            for val in non_empty_values[:10]:  # Check first 10 values
                try:
                    # Check if it's a number
                    float(val)
                    numeric_count += 1
                except:
                    # Check if it's a date pattern
                    if '-' in val and len(val) == 10:  # Likely YYYY-MM-DD
                        numeric_count += 1

            # If more than 30% of values are numeric/dates, likely a data row
            if numeric_count > len(non_empty_values[:10]) * 0.3:
                print(f"DEBUG: Row {row_num} looks like data (too many numbers/dates), skipping...")
                continue

            # Check if this row contains header keywords
            if len(non_empty_values) >= 5:  # Real headers should have multiple columns
                row_text = ' '.join(non_empty_values).lower()
                matches = sum(1 for keyword in header_keywords if keyword.lower() in row_text)

                # Give extra weight to specific critical keywords
                if 'l2' in row_text and ('id' in row_text or '编号' in row_text):
                    matches += 3  # L2 ID is a critical field
                if 'l2id' in row_text.replace(' ', '').lower():
                    matches += 3  # L2ID without space
                if '应用' in row_text or 'application' in row_text.lower():
                    matches += 2
                if '负责' in row_text or '团队' in row_text:
                    matches += 2
                if '【计划】' in row_text or '【实际】' in row_text:
                    matches += 3  # Strong indicator of header row
                if '监管' in row_text and '验收' in row_text:
                    matches += 2

                # Calculate score based on matches and non-empty cells
                score = matches * 10 + len(non_empty_values)

                if score > best_match['score']:
                    best_match = {
                        'row': row_num,
                        'score': score,
                        'headers': [str(cell.value).strip() if cell.value else '' for cell in row]
                    }
                    print(f"DEBUG: Row {row_num} - score: {score}, matches: {matches}, non-empty: {len(non_empty_values)}")

        # Use the best matching row as headers
        if best_match['score'] > 20:  # Minimum score threshold
            header_row = best_match['row']
            headers = best_match['headers']
            print(f"DEBUG: Selected header row {header_row} with score {best_match['score']}")
        else:
            print(f"DEBUG: Score too low ({best_match['score']}), trying fallback detection...")

            # Fallback: if no headers found with keywords, look for row with many non-empty cells
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=30), start=1):
                row_values = [cell.value for cell in row]
                non_empty_values = [str(v).strip() for v in row_values if v is not None and str(v).strip()]

                # Skip statistics rows
                if non_empty_values:
                    first_cell_text = non_empty_values[0].lower()
                    if any(skip_word in first_cell_text for skip_word in skip_keywords):
                        continue

                # Look for rows with many short, distinct values (typical of headers)
                if len(non_empty_values) >= 8:
                    # Check if values are reasonably short and diverse
                    avg_length = sum(len(v) for v in non_empty_values) / len(non_empty_values)
                    unique_ratio = len(set(non_empty_values)) / len(non_empty_values)

                    if avg_length < 30 and unique_ratio > 0.7:  # Short and diverse = likely headers
                        headers = [str(cell.value).strip() if cell.value else '' for cell in row]
                        header_row = row_num
                        print(f"DEBUG: Using fallback header detection at row {header_row} (avg_len: {avg_length:.1f}, unique: {unique_ratio:.2f})")
                        break

            # If still no headers found, use the best match we had
            if not headers and best_match['headers']:
                header_row = best_match['row']
                headers = best_match['headers']
                print(f"DEBUG: Using best available match at row {header_row}")

        print(f"DEBUG: Found headers at row {header_row}: {headers}")  # 显示所有标题
        print(f"DEBUG: Available field mappings: {list(field_mapping.keys())[:20]}")  # 显示前20个可用映射

        # Map headers to database fields
        column_mapping = {}
        unmapped_headers = []
        for i, header in enumerate(headers):
            # Clean header - remove extra spaces and newlines
            header_cleaned = header.strip().replace('\n', ' ').replace('  ', ' ')

            # Try exact match first
            if header in field_mapping:
                column_mapping[i] = field_mapping[header]
                print(f"DEBUG: [MAPPED EXACT] column {i}: '{header}' -> '{field_mapping[header]}'")  # 调试信息
            # Try cleaned version
            elif header_cleaned in field_mapping:
                column_mapping[i] = field_mapping[header_cleaned]
                print(f"DEBUG: [MAPPED CLEANED] column {i}: '{header}' (cleaned: '{header_cleaned}') -> '{field_mapping[header_cleaned]}'")
            # Special handling for headers with 【】markers which might have different formatting
            elif '【' in header:
                # Try removing all spaces around 【】
                header_normalized = header.replace(' ', '').replace('\n', '')
                for key in field_mapping:
                    if '【' in key and key.replace(' ', '').replace('\n', '') == header_normalized:
                        column_mapping[i] = field_mapping[key]
                        print(f"DEBUG: [MAPPED NORMALIZED] column {i}: '{header}' -> '{field_mapping[key]}'")
                        break
                else:
                    unmapped_headers.append(header)
            else:
                unmapped_headers.append(header)

        print(f"DEBUG: Column mapping: {column_mapping}")  # 调试信息
        print(f"DEBUG: [UNMAPPED] headers: {unmapped_headers}")  # 显示未映射的标题

        # Check if we have at least the minimum required mappings
        if 'l2_id' not in column_mapping.values() and 'application_l2_id' not in column_mapping.values():
            print("DEBUG: [WARNING] No L2 ID column found in Excel file")

        # 如果没有映射到任何列，尝试智能推断
        if not column_mapping:
            print("DEBUG: [INFO] No direct column mapping found, trying intelligent matching...")

            # 尝试模糊匹配常见的字段名（支持Applications和SubTasks）
            fuzzy_mapping = {
                # Application相关的字段模式
                'l2': 'l2_id',
                'L2': 'l2_id',
                '序号': 'l2_id',
                '编号': 'l2_id',
                '应用': 'app_name',
                '系统': 'app_name',
                '名称': 'app_name',
                '监管': 'supervision_year',
                '年度': 'supervision_year',
                '指标': 'supervision_year',
                '改造': 'transformation_target',
                '目标': 'transformation_target',
                'AK': 'transformation_target',
                '云原生': 'transformation_target',
                '阶段': 'current_stage',
                '当前': 'current_stage',
                '状态': 'overall_status',
                '整体': 'overall_status',
                '总体': 'overall_status',
                '团队': 'responsible_team',
                '部门': 'responsible_team',
                '负责人': 'responsible_person',
                '责任人': 'responsible_person',
                '进度': 'progress_percentage',
                '百分比': 'progress_percentage',
                'progress': 'progress_percentage',
                'status': 'overall_status',
                'team': 'responsible_team',
                'person': 'responsible_person',
                '计划': 'planned_biz_online_date',
                '实际': 'actual_biz_online_date',
                '上线': 'planned_biz_online_date',
                '完成': 'planned_biz_online_date',
                '备注': 'notes',
                '说明': 'notes',
                'note': 'notes',

                # SubTask相关的字段模式
                '模块': 'module_name',
                'module': 'module_name',
                'blocked': 'is_blocked',
                '阻塞': 'is_blocked',
                'assign': 'assigned_to',
                '分配': 'assigned_to'
            }

            for i, header in enumerate(headers):
                header_lower = header.lower().strip()
                for pattern, field in fuzzy_mapping.items():
                    if pattern in header_lower:
                        column_mapping[i] = field
                        print(f"DEBUG: [FUZZY] matched '{header}' -> '{field}' (pattern: '{pattern}')")
                        break

        # Extract data rows with smart termination and chunked processing
        data = []
        row_count = 0
        empty_row_count = 0
        chunk_size = 500  # Smaller chunks for better performance
        max_rows_to_process = 50000  # Reasonable limit to prevent timeout
        max_consecutive_empty = 20  # Stop after 20 consecutive empty rows

        # Get actual data range (not worksheet.max_row which can be misleading)
        actual_max_row = min(worksheet.max_row, header_row + max_rows_to_process)

        # For very large files, try to find the actual end of data
        if worksheet.max_row > 10000:
            print(f"DEBUG: Large file detected ({worksheet.max_row} rows), finding actual data range...")
            # Sample check: if rows 1000-1100 are all empty, likely end of data
            sample_start = min(1000, worksheet.max_row)
            sample_empty = 0
            for row in worksheet.iter_rows(min_row=sample_start, max_row=sample_start + 100):
                if all(cell.value is None for cell in row):
                    sample_empty += 1

            if sample_empty > 90:  # If >90% of sample is empty, data likely ends before
                actual_max_row = sample_start
                print(f"DEBUG: Detected probable end of data around row {actual_max_row}")

        total_rows = actual_max_row - header_row
        print(f"DEBUG: Processing up to {total_rows} rows from Excel file...")

        # Process rows in chunks for better performance
        for chunk_start in range(header_row + 1, actual_max_row + 1, chunk_size):
            chunk_end = min(chunk_start + chunk_size - 1, actual_max_row)
            chunk_data = []
            chunk_empty_count = 0

            for row in worksheet.iter_rows(min_row=chunk_start, max_row=chunk_end, values_only=True):
                # Check if row is completely empty
                if all(v is None for v in row):
                    empty_row_count += 1
                    chunk_empty_count += 1

                    # Stop if too many consecutive empty rows
                    if empty_row_count >= max_consecutive_empty:
                        print(f"DEBUG: Found {empty_row_count} consecutive empty rows, assuming end of data at row {chunk_start + chunk_empty_count}")
                        break
                    continue
                else:
                    empty_row_count = 0  # Reset counter when we find data

                row_data = {}
                has_data = False

                for i, value in enumerate(row):
                    if i in column_mapping:
                        field_name = column_mapping[i]

                        # Convert data types
                        if value is not None:
                            value = self._convert_cell_value(value, field_name)
                            has_data = True

                        row_data[field_name] = value

                if has_data:
                    chunk_data.append(row_data)
                    row_count += 1

                    # Only print first few rows for debugging
                    if row_count <= 3:
                        print(f"DEBUG: Row {chunk_start + row_count - 1} data: {row_data}")

            # Add chunk data to main data list
            data.extend(chunk_data)

            # Stop if we found the end of data
            if empty_row_count >= max_consecutive_empty:
                break

            # Stop if we've processed enough rows
            if row_count >= max_rows_to_process:
                print(f"DEBUG: Reached maximum row limit ({max_rows_to_process}), stopping processing")
                break

            # Progress indicator for large files
            if row_count > 0 and row_count % 1000 == 0:
                print(f"DEBUG: Processed {row_count} rows...")

        print(f"DEBUG: Total rows extracted: {len(data)}")  # 调试信息

        # Convert to DataFrame and optimize memory usage
        if data:
            df = pd.DataFrame(data)

            # Enhanced debugging to see what's in the DataFrame
            print(f"DEBUG: DataFrame columns after mapping: {list(df.columns)}")
            print(f"DEBUG: DataFrame shape: {df.shape}")
            if len(df) > 0:
                print(f"DEBUG: First row data sample:")
                for col in df.columns:
                    if col in ['l2_id', 'app_name', 'dev_team', 'overall_transformation_target', 'current_status']:
                        print(f"  {col}: {df.iloc[0].get(col, 'N/A')}")

                # Check specifically for app_name column
                if 'app_name' in df.columns:
                    non_empty_app_names = df[df['app_name'].notna() & (df['app_name'] != '')]
                    print(f"DEBUG: Found {len(non_empty_app_names)} rows with non-empty app_name values")
                    if len(non_empty_app_names) > 0:
                        print(f"DEBUG: Sample app_names: {list(non_empty_app_names['app_name'].head(3))}")
                else:
                    print(f"DEBUG: WARNING - app_name column not found in DataFrame!")

            # Optimize memory for large dataframes
            for col in df.columns:
                col_type = df[col].dtype
                if col_type != 'object':
                    try:
                        df[col] = pd.to_numeric(df[col], downcast='integer')
                    except:
                        pass
            return df
        else:
            return pd.DataFrame()

    def _convert_cell_value(self, value: Any, field_name: str) -> Any:
        """Convert cell value to appropriate Python type."""

        if value is None or value == '':
            return None

        try:
            # Special handling for l2_id - always treat as string
            if field_name == 'l2_id':
                if isinstance(value, (int, float)):
                    import math
                    if math.isnan(value):
                        return None
                    return str(int(value))  # Convert numeric to string
                else:
                    return str(value).strip() if value else None
            # DateTime fields (timestamp fields)
            if field_name in self.config.DATETIME_FIELDS:
                if isinstance(value, datetime):
                    return value
                elif isinstance(value, date):
                    # Convert date to datetime at midnight
                    return datetime.combine(value, datetime.min.time())
                elif isinstance(value, (int, float)):
                    # Handle Excel numeric date format
                    import math
                    if math.isnan(value):
                        return None
                    try:
                        # Excel stores dates as days since 1900-01-01
                        excel_base_date = datetime(1899, 12, 30)
                        return excel_base_date + timedelta(days=int(value))
                    except (ValueError, OverflowError):
                        return None
                elif isinstance(value, str):
                    # For DateTime fields, if it's a status string like '已完成', return None
                    # These strings should be mapped to other fields
                    status_strings = ['已完成', '通过', '检查通过', '未完成', '进行中', '待检查']
                    if value.strip() in status_strings:
                        return None
                    # Try to parse datetime string
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
                        try:
                            return datetime.strptime(value.strip(), fmt)
                        except ValueError:
                            continue
                    return None

            # Date fields
            elif field_name in self.config.DATE_FIELDS:
                if isinstance(value, datetime):
                    return value.date()
                elif isinstance(value, date):
                    return value
                elif isinstance(value, (int, float)):
                    # Handle Excel numeric date format
                    import math
                    if math.isnan(value):
                        return None
                    try:
                        # Excel stores dates as days since 1900-01-01 (with a leap year bug)
                        # For dates after 1900-03-01, we need to subtract 2
                        # (1 for the bug, 1 for 0-indexing)
                        excel_base_date = datetime(1899, 12, 30)  # Excel's base date accounting for the bug
                        return (excel_base_date + timedelta(days=int(value))).date()
                    except (ValueError, OverflowError):
                        return None
                elif isinstance(value, str):
                    # Try to parse date string
                    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                        try:
                            return datetime.strptime(value.strip(), fmt).date()
                        except ValueError:
                            continue
                    return None

            # Integer fields
            elif field_name in self.config.INTEGER_FIELDS:
                if isinstance(value, (int, float)):
                    # Check for NaN values
                    import math
                    if math.isnan(value):
                        return None
                    return int(value)
                elif isinstance(value, str):
                    # Special handling for certain fields
                    if field_name == 'ak_supervision_acceptance_year':
                        # Extract year from strings like "2025年"
                        import re
                        # 处理datetime对象
                        if isinstance(value, datetime):
                            return value.year
                        # 处理date对象
                        if isinstance(value, date):
                            return value.year
                        # 处理字符串类型
                        year_match = re.search(r'(\d{4})', str(value))
                        if year_match:
                            return int(year_match.group(1))
                    elif field_name == 'app_tier':
                        # Convert tier strings to numbers
                        tier_mapping = {
                            '第一级': 1, '第1级': 1, '一级': 1, '1级': 1,
                            '第二级': 2, '第2级': 2, '二级': 2, '2级': 2,
                            '第三级': 3, '第3级': 3, '三级': 3, '3级': 3,
                            '第四级': 4, '第4级': 4, '四级': 4, '4级': 4,
                            '第五级': 5, '第5级': 5, '五级': 5, '5级': 5
                        }
                        if value.strip() in tier_mapping:
                            return tier_mapping[value.strip()]
                    # Try standard conversion
                    try:
                        return int(float(value.strip()))
                    except:
                        return None

            # Boolean fields
            elif field_name in self.config.BOOLEAN_FIELDS:
                if isinstance(value, bool):
                    return value
                elif isinstance(value, str):
                    value_lower = value.strip().lower()
                    return value_lower in ['是', 'true', 'yes', '1', 'y']
                elif isinstance(value, (int, float)):
                    return bool(value)

            # String fields (explicitly handle known string fields)
            elif field_name in self.config.STRING_FIELDS:
                return str(value).strip() if value else None

            # Default handling for other fields
            else:
                return str(value).strip() if value else None

        except (ValueError, TypeError):
            return str(value).strip() if value else None

        return value

    async def _validate_applications_data(self, db: AsyncSession, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Validate applications data."""

        errors = []

        for index, row in df.iterrows():
            row_num = index + 2  # Excel row number (1-based + header)

            # Check required fields
            for field in self.config.APPLICATION_REQUIRED:
                if pd.isna(row.get(field)) or row.get(field) == '' or row.get(field) is None:
                    errors.append({
                        'row': row_num,
                        'column': self._get_column_name(field, self.config.APPLICATION_FIELDS),
                        'message': f'必填字段不能为空: {field}',
                        'value': row.get(field)
                    })

            # Validate L2 ID format (but don't add prefix)
            l2_id = row.get('l2_id')
            if l2_id:
                l2_id_str = str(l2_id).strip()
                # Keep the original L2 ID without adding prefix
                # User explicitly requested: "对于导入的数据，请不要在原数据前加前缀"
                df.at[index, 'l2_id'] = l2_id_str

            # Convert supervision year if it's a string
            year = row.get('ak_supervision_acceptance_year')
            if year:
                # Convert if it's a string like "2025年"
                if isinstance(year, str):
                    import re
                    year_match = re.search(r'(\d{4})', year)
                    if year_match:
                        year = int(year_match.group(1))
                        df.at[index, 'ak_supervision_acceptance_year'] = year
                # Note: Removed year range validation as per user request
                # Years can be any valid value

            # Validate transformation target (支持前端发送的值)
            target = row.get('overall_transformation_target')
            if target:
                # 标准化转型目标值
                target_mapping = {
                    'cloud_native': TransformationTarget.CLOUD_NATIVE.value,
                    'AK': TransformationTarget.AK.value,
                    'ak': TransformationTarget.AK.value,
                    '云原生': TransformationTarget.CLOUD_NATIVE.value,
                    'Cloud Native': TransformationTarget.CLOUD_NATIVE.value
                }

                if target in target_mapping:
                    df.at[index, 'overall_transformation_target'] = target_mapping[target]
                elif target not in [t.value for t in TransformationTarget]:
                    # 如果不匹配，使用默认值
                    df.at[index, 'overall_transformation_target'] = TransformationTarget.AK.value

            # Validate and normalize status
            status = row.get('current_status')
            if status:
                # 状态映射 - 根据实际需求
                status_mapping = {
                    # 保持原状态
                    '待启动': '待启动',
                    '需求进行中': '需求进行中',
                    '研发进行中': '研发进行中',
                    '业务上线中': '业务上线中',
                    '阻塞': '阻塞',
                    '全部完成': '全部完成',

                    # 需要映射的状态
                    '部署进行中': '技术上线中',
                    '中止': '计划下线',

                    # 英文映射（如果前端使用）
                    'not_started': '待启动',
                    'requirement_in_progress': '需求进行中',
                    'dev_in_progress': '研发进行中',
                    'in_progress': '研发进行中',
                    'tech_online': '技术上线中',
                    'deployment_in_progress': '技术上线中',
                    'biz_online': '业务上线中',
                    'blocked': '阻塞',
                    'planned_offline': '计划下线',
                    'terminated': '计划下线',
                    'cancelled': '计划下线',
                    'completed': '全部完成',

                    # 其他可能的变体
                    '正常': '研发进行中',
                    '进行中': '研发进行中',
                    '完成': '全部完成',
                    '已完成': '全部完成',
                    '未开始': '待启动',
                    '测试中': '技术上线中',
                    '待部署': '技术上线中',
                    '已上线': '全部完成',
                    '已阻塞': '阻塞',
                    '已中止': '计划下线',
                    '已取消': '计划下线'
                }

                if status in status_mapping:
                    df.at[index, 'current_status'] = status_mapping[status]
                elif status not in ['待启动', '需求进行中', '研发进行中', '技术上线中',
                                    '业务上线中', '阻塞', '计划下线', '全部完成']:
                    # 默认值
                    df.at[index, 'current_status'] = '待启动'
            else:
                # 状态为空时设置默认值
                df.at[index, 'current_status'] = '待启动'

            # Convert app_tier if it's a string
            app_tier = row.get('app_tier')
            if app_tier and isinstance(app_tier, str):
                tier_mapping = {
                    '第一级': 1, '第1级': 1, '一级': 1, '1级': 1,
                    '第二级': 2, '第2级': 2, '二级': 2, '2级': 2,
                    '第三级': 3, '第3级': 3, '三级': 3, '3级': 3,
                    '第四级': 4, '第4级': 4, '四级': 4, '4级': 4,
                    '第五级': 5, '第5级': 5, '五级': 5, '5级': 5
                }
                if app_tier.strip() in tier_mapping:
                    df.at[index, 'app_tier'] = tier_mapping[app_tier.strip()]

            # Validate progress percentage
            progress = row.get('progress_percentage')
            if progress is not None and (progress < 0 or progress > 100):
                errors.append({
                    'row': row_num,
                    'column': '进度百分比',
                    'message': '进度百分比必须在0-100之间',
                    'value': progress
                })

        # Check for duplicate L2 IDs
        l2_ids = df['l2_id'].dropna()
        duplicates = l2_ids[l2_ids.duplicated()].tolist()
        if duplicates:
            for dup_id in duplicates:
                dup_rows = df[df['l2_id'] == dup_id].index + 2
                for row_num in dup_rows:
                    errors.append({
                        'row': row_num,
                        'column': 'L2 ID',
                        'message': f'L2 ID重复: {dup_id}',
                        'value': dup_id
                    })

        return errors

    async def _validate_subtasks_data(self, db: AsyncSession, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Validate subtasks data based on actual database structure."""

        errors = []
        warnings = []

        # Get valid application mappings (l2_id -> database id)
        app_result = await db.execute(select(Application.id, Application.l2_id))
        app_mapping = {l2_id: app_id for app_id, l2_id in app_result.all()}
        valid_l2_ids = set(app_mapping.keys())

        print(f"DEBUG: Found {len(valid_l2_ids)} valid applications in database")

        # Count statistics
        empty_l2_count = 0
        rows_to_process = 0

        for index, row in df.iterrows():
            row_num = index + 2  # Excel row number (1-based + header)

            # Check L2 ID (this is the Application's l2_id string, not the database id)
            app_l2_id = row.get('l2_id')

            # Skip rows with empty L2 ID
            if pd.isna(app_l2_id) or app_l2_id == '' or app_l2_id is None:
                empty_l2_count += 1
                continue  # Skip this row silently

            rows_to_process += 1
            app_l2_id_str = str(app_l2_id).strip()

            # Store cleaned L2 ID
            df.at[index, 'l2_id'] = app_l2_id_str

            # Check if application exists (will be created if not)
            if app_l2_id_str not in valid_l2_ids:
                warnings.append({
                    'row': row_num,
                    'type': 'info',
                    'message': f'应用 {app_l2_id_str} 不存在，将自动创建'
                })

            # Validate and normalize sub_target
            sub_target = row.get('sub_target')
            if sub_target:
                # Standardize sub_target values
                target_mapping = {
                    'cloud_native': '云原生',
                    'Cloud Native': '云原生',
                    'CLOUD_NATIVE': '云原生',
                    '云原生': '云原生',
                    'AK': 'AK',
                    'ak': 'AK',
                    'Ak': 'AK'
                }

                if sub_target in target_mapping:
                    df.at[index, 'sub_target'] = target_mapping[sub_target]
                elif sub_target not in ['AK', '云原生']:
                    # Default to AK if invalid
                    df.at[index, 'sub_target'] = 'AK'
                    warnings.append({
                        'row': row_num,
                        'column': '子目标',
                        'message': f'无效的子目标 "{sub_target}"，已设为默认值 "AK"'
                    })

            # Validate and normalize task_status
            task_status = row.get('task_status')
            if task_status:
                # 状态映射 - 根据实际需求更新
                status_mapping = {
                    # 保持原状态
                    '未开始': '未开始',
                    '需求进行中': '需求进行中',
                    '研发进行中': '研发进行中',
                    '业务上线中': '业务上线中',
                    '阻塞': '阻塞',
                    '子任务完成': '子任务完成',

                    # 需要映射的状态
                    '部署进行中': '技术上线中',
                    '中止': '计划下线',

                    # 英文映射（如果前端有使用）
                    'not_started': '未开始',
                    'NOT_STARTED': '未开始',
                    'requirement_in_progress': '需求进行中',
                    'REQUIREMENT_IN_PROGRESS': '需求进行中',
                    'dev_in_progress': '研发进行中',
                    'DEV_IN_PROGRESS': '研发进行中',
                    'in_progress': '研发进行中',
                    'tech_online': '技术上线中',
                    'TECH_ONLINE': '技术上线中',
                    'deployment_in_progress': '技术上线中',
                    'biz_online': '业务上线中',
                    'BIZ_ONLINE': '业务上线中',
                    'blocked': '阻塞',
                    'BLOCKED': '阻塞',
                    'planned_offline': '计划下线',
                    'PLANNED_OFFLINE': '计划下线',
                    'terminated': '计划下线',
                    'cancelled': '计划下线',
                    'completed': '子任务完成',
                    'COMPLETED': '子任务完成',

                    # 其他可能的变体
                    '进行中': '研发进行中',
                    '完成': '子任务完成',
                    '已完成': '子任务完成',
                    '待启动': '未开始',
                    '测试中': '技术上线中',
                    '待部署': '技术上线中',
                    '已上线': '子任务完成',
                    '已阻塞': '阻塞',
                    '已中止': '计划下线',
                    '已取消': '计划下线'
                }

                if task_status in status_mapping:
                    df.at[index, 'task_status'] = status_mapping[task_status]
                elif task_status not in ['未开始', '需求进行中', '研发进行中', '技术上线中',
                                         '业务上线中', '阻塞', '计划下线', '子任务完成']:
                    # 如果不在有效状态列表中，默认设为"未开始"
                    df.at[index, 'task_status'] = '未开始'
                    warnings.append({
                        'row': row_num,
                        'column': '任务状态',
                        'message': f'无效的状态 "{task_status}"，已设为默认值 "未开始"'
                    })
            else:
                # 如果状态为空，设置默认值
                df.at[index, 'task_status'] = '未开始'

            # 根据状态设置默认进度百分比
            progress = row.get('progress_percentage')
            if pd.isna(progress) or progress is None:
                # 根据状态设置默认进度
                status_progress_map = {
                    '未开始': 0,
                    '需求进行中': 10,
                    '研发进行中': 30,
                    '技术上线中': 60,
                    '业务上线中': 80,
                    '子任务完成': 100,
                    '阻塞': None,  # 保持原有进度
                    '计划下线': None  # 保持原有进度
                }

                current_status = df.at[index, 'task_status']
                default_progress = status_progress_map.get(current_status, 0)
                if default_progress is not None:
                    df.at[index, 'progress_percentage'] = default_progress

            # Validate boolean fields
            for bool_field in ['is_blocked', 'resource_applied']:
                value = row.get(bool_field)
                if value is not None:
                    if isinstance(value, str):
                        value_lower = value.strip().lower()
                        df.at[index, bool_field] = value_lower in ['是', 'true', 'yes', '1', 'y', '已申请', '已阻塞']
                    elif isinstance(value, (int, float)):
                        df.at[index, bool_field] = bool(value)
                    elif pd.isna(value):
                        df.at[index, bool_field] = False
                else:
                    df.at[index, bool_field] = False  # Default to False

            # Validate date fields - handle NaT values
            for date_field in ['planned_requirement_date', 'planned_release_date',
                               'planned_tech_online_date', 'planned_biz_online_date',
                               'actual_requirement_date', 'actual_release_date',
                               'actual_tech_online_date', 'actual_biz_online_date']:
                date_value = row.get(date_field)
                if pd.isna(date_value):
                    df.at[index, date_field] = None

            # Validate ops_requirement_submitted (DateTime field)
            ops_req = row.get('ops_requirement_submitted')
            if ops_req is not None and not pd.isna(ops_req):
                # Check if it's a string that looks like a status rather than a date
                if isinstance(ops_req, str):
                    status_strings = ['已完成', '已提交', '通过', '未提交', '进行中', '待提交', '完成']
                    if ops_req.strip() in status_strings:
                        # This is likely a mismatched column, clear it
                        df.at[index, 'ops_requirement_submitted'] = None
            else:
                df.at[index, 'ops_requirement_submitted'] = None

            # Validate ops_testing_status and launch_check_status (these are strings)
            for status_field in ['ops_testing_status', 'launch_check_status']:
                status_value = row.get(status_field)
                if status_value and isinstance(status_value, str):
                    # Normalize common status values
                    status_normalize = {
                        '通过': '通过',
                        '已通过': '通过',
                        '完成': '通过',
                        '已完成': '通过',
                        '检查通过': '通过',
                        '未完成': '进行中',
                        '进行中': '进行中',
                        '待检查': '待检查',
                        '未开始': '待检查'
                    }
                    if status_value in status_normalize:
                        df.at[index, status_field] = status_normalize[status_value]
                elif pd.isna(status_value):
                    df.at[index, status_field] = None

        # Summary messages
        if empty_l2_count > 0:
            warnings.append({
                'type': 'summary',
                'message': f'共 {empty_l2_count} 行因L2 ID为空被跳过'
            })

        if rows_to_process == 0:
            errors.append({
                'type': 'critical',
                'message': '没有找到有效的数据行（所有行的L2 ID都为空）'
            })

        # Check for duplicate entries (same l2_id + sub_target + version_name)
        if rows_to_process > 0:
            # Only check duplicates for non-null combinations
            df_valid = df[df['l2_id'].notna()].copy()

            # Handle null version_name by replacing with empty string for grouping
            df_valid['version_name'] = df_valid['version_name'].fillna('')

            duplicates = df_valid.groupby(['l2_id', 'sub_target', 'version_name']).size()
            duplicates = duplicates[duplicates > 1]

            for (l2_id, sub_target, version_name), count in duplicates.items():
                # Find duplicate rows
                mask = (df['l2_id'] == l2_id) & (df['sub_target'] == sub_target)
                if version_name:
                    mask = mask & (df['version_name'] == version_name)
                else:
                    mask = mask & df['version_name'].isna()

                duplicate_rows = df[mask].index + 2

                for row_num in duplicate_rows:
                    errors.append({
                        'row': row_num,
                        'column': 'L2 ID + 子目标 + 版本名',
                        'message': f'重复的子任务: L2={l2_id}, 目标={sub_target}, 版本={version_name or "空"}',
                        'value': f'{l2_id}-{sub_target}-{version_name}'
                    })

        # Log validation summary
        print(
            f"DEBUG: Validation completed - Errors: {len(errors)}, Warnings: {len(warnings)}, Rows to process: {rows_to_process}")

        return errors  # Return only errors, warnings are logged but don't block import

    def _get_column_name(self, field_name: str, field_mapping: Dict[str, str]) -> str:
        """Get Excel column name from field name."""
        for column_name, mapped_field in field_mapping.items():
            if mapped_field == field_name:
                return column_name
        return field_name

    async def _import_applications_data(self, db: AsyncSession, df: pd.DataFrame, user: User) -> Dict[str, int]:
        """Import applications data to database."""

        imported = 0
        updated = 0
        skipped = 0

        # Get user_id as integer to avoid relationship loading
        user_id = user.id if user else None
        if not user_id:
            raise ValueError("User ID is required for import")

        # Store original autoflush state and disable it
        original_autoflush = db.autoflush
        db.autoflush = False

        # Collections for batch operations
        new_applications = []
        updated_applications = []

        try:
            # First, get all existing applications in one query
            existing_l2_ids = df['l2_id'].dropna().unique().tolist()
            if existing_l2_ids:
                stmt = select(Application).where(Application.l2_id.in_(existing_l2_ids))
                result = await db.execute(stmt)
                existing_apps_map = {app.l2_id: app for app in result.scalars()}
            else:
                existing_apps_map = {}

            for index, row in df.iterrows():
                try:
                    l2_id = row.get('l2_id')
                    if not l2_id:
                        skipped += 1
                        continue

                    existing_app = existing_apps_map.get(l2_id)

                    if existing_app:
                        # Update existing application
                        application_model_fields = {
                            'l2_id', 'app_name', 'ak_supervision_acceptance_year', 'overall_transformation_target',
                            'current_transformation_phase', 'current_status', 'app_tier', 'belonging_l1_name',
                            'belonging_projects', 'is_ak_completed', 'is_cloud_native_completed',
                            'is_domain_transformation_completed', 'is_dbpm_transformation_completed',
                            'dev_mode', 'ops_mode', 'dev_owner', 'dev_team', 'ops_owner', 'ops_team',
                            'belonging_kpi', 'acceptance_status', 'planned_requirement_date', 'planned_release_date',
                            'planned_tech_online_date', 'planned_biz_online_date', 'actual_requirement_date',
                            'actual_release_date', 'actual_tech_online_date', 'actual_biz_online_date',
                            'is_delayed', 'delay_days', 'notes'
                        }

                        for field, value in row.items():
                            if field in application_model_fields and value is not None and value != '' and hasattr(existing_app, field):
                                try:
                                    # Handle NaN values
                                    if pd.isna(value):
                                        continue  # Skip NaN values
                                    
                                    # Handle date fields
                                    if field in self.config.DATE_FIELDS:
                                        if isinstance(value, str):
                                            # Try to parse date string
                                            try:
                                                value = datetime.strptime(value, '%Y-%m-%d').date()
                                            except:
                                                try:
                                                    value = datetime.strptime(value, '%Y/%m/%d').date()
                                                except:
                                                    print(f"[WARNING] Could not parse date '{value}' for field {field}, skipping")
                                                    continue
                                        elif isinstance(value, datetime):
                                            value = value.date()
                                    
                                    # Handle boolean fields
                                    elif field in self.config.BOOLEAN_FIELDS:
                                        if isinstance(value, str):
                                            value = value.lower() in ['true', 'yes', '是', '1', 't', 'y']
                                    
                                    # Handle integer fields
                                    elif field in self.config.INTEGER_FIELDS:
                                        if pd.notna(value):
                                            # Special handling for year field which might be a datetime
                                            if field == 'ak_supervision_acceptance_year':
                                                if isinstance(value, datetime):
                                                    value = value.year
                                                elif isinstance(value, date):
                                                    value = value.year
                                                elif isinstance(value, str):
                                                    # Extract year from string like "2025年" or "2025"
                                                    import re
                                                    year_match = re.search(r'(\d{4})', value)
                                                    if year_match:
                                                        value = int(year_match.group(1))
                                                    else:
                                                        print(f"[WARNING] Could not extract year from '{value}', skipping")
                                                        continue
                                                else:
                                                    value = int(float(value))  # Normal numeric conversion
                                            else:
                                                value = int(float(value))  # Convert through float to handle decimals
                                    
                                    setattr(existing_app, field, value)
                                except Exception as e:
                                    print(f"[WARNING] Error setting field {field} with value '{value}': {e}")
                                    continue

                        existing_app.updated_by = user_id
                        updated_applications.append(existing_app)
                        updated += 1
                    else:
                        # Create new application
                        app_data = {}
                        application_model_fields = {
                            'l2_id', 'app_name', 'ak_supervision_acceptance_year', 'overall_transformation_target',
                            'current_transformation_phase', 'current_status', 'app_tier', 'belonging_l1_name',
                            'belonging_projects', 'is_ak_completed', 'is_cloud_native_completed',
                            'is_domain_transformation_completed', 'is_dbpm_transformation_completed',
                            'dev_mode', 'ops_mode', 'dev_owner', 'dev_team', 'ops_owner', 'ops_team',
                            'belonging_kpi', 'acceptance_status', 'planned_requirement_date', 'planned_release_date',
                            'planned_tech_online_date', 'planned_biz_online_date', 'actual_requirement_date',
                            'actual_release_date', 'actual_tech_online_date', 'actual_biz_online_date',
                            'is_delayed', 'delay_days', 'notes'
                        }

                        for k, v in row.items():
                            if k in application_model_fields and v is not None and v != '':
                                try:
                                    # Handle NaN values
                                    if pd.isna(v):
                                        continue  # Skip NaN values
                                    
                                    # Handle date fields
                                    if k in self.config.DATE_FIELDS:
                                        if isinstance(v, str):
                                            try:
                                                v = datetime.strptime(v, '%Y-%m-%d').date()
                                            except:
                                                try:
                                                    v = datetime.strptime(v, '%Y/%m/%d').date()
                                                except:
                                                    print(f"[WARNING] Could not parse date '{v}' for field {k}, skipping")
                                                    continue
                                        elif isinstance(v, datetime):
                                            v = v.date()
                                    
                                    # Handle boolean fields
                                    elif k in self.config.BOOLEAN_FIELDS:
                                        if isinstance(v, str):
                                            v = v.lower() in ['true', 'yes', '是', '1', 't', 'y']
                                    
                                    # Handle integer fields
                                    elif k in self.config.INTEGER_FIELDS:
                                        if pd.notna(v):
                                            # Special handling for year field which might be a datetime
                                            if k == 'ak_supervision_acceptance_year':
                                                if isinstance(v, datetime):
                                                    v = v.year
                                                elif isinstance(v, date):
                                                    v = v.year
                                                elif isinstance(v, str):
                                                    # Extract year from string like "2025年" or "2025"
                                                    import re
                                                    year_match = re.search(r'(\d{4})', v)
                                                    if year_match:
                                                        v = int(year_match.group(1))
                                                    else:
                                                        print(f"[WARNING] Could not extract year from '{v}', skipping")
                                                        continue
                                                else:
                                                    v = int(float(v))  # Normal numeric conversion
                                            else:
                                                v = int(float(v))  # Convert through float to handle decimals
                                    
                                    app_data[k] = v
                                except Exception as e:
                                    print(f"[WARNING] Error processing field {k} with value '{v}': {e}")
                                    continue

                        app_data['created_by'] = user_id
                        app_data['updated_by'] = user_id

                        # Set defaults
                        if 'app_name' not in app_data or not app_data['app_name']:
                            app_data['app_name'] = '未命名应用'
                        if 'overall_transformation_target' not in app_data or not app_data['overall_transformation_target']:
                            app_data['overall_transformation_target'] = 'AK'
                        if 'current_status' not in app_data or not app_data['current_status']:
                            app_data['current_status'] = '待启动'

                        new_app = Application(**app_data)
                        new_applications.append(new_app)
                        imported += 1

                except Exception as e:
                    print(f"[ERROR] Failed processing application row {index + 2} (L2_ID: {row.get('l2_id', 'unknown')}): {e}")
                    import traceback
                    traceback.print_exc()
                    skipped += 1
                    continue

            # Add all new applications in batch
            if new_applications:
                print(f"[INFO] Adding {len(new_applications)} new applications to session...")
                for app in new_applications:
                    try:
                        db.add(app)
                    except Exception as e:
                        print(f"[ERROR] Failed to add application {app.l2_id}: {e}")
                        raise
                print(f"[INFO] All new applications added to session successfully")

            # Commit all changes at once
            print(f"[INFO] Committing {imported} new and {updated} updated applications...")
            await db.commit()
            print(f"[INFO] Successfully committed all application changes")

        except Exception as e:
            # Restore autoflush and rollback on error
            db.autoflush = original_autoflush
            await db.rollback()
            print(f"[ERROR] Failed to import applications: {e}")
            import traceback
            traceback.print_exc()
            raise

        finally:
            # Always restore autoflush state
            db.autoflush = original_autoflush

        # Log final statistics
        print(f"[INFO] Import statistics - Imported: {imported}, Updated: {updated}, Skipped: {skipped}")
        if skipped > 0:
            print(f"[WARNING] {skipped} rows were skipped during import")

        return {
            'imported': imported,
            'updated': updated,
            'skipped': skipped
        }

    async def _import_subtasks_data(self, db: AsyncSession, df: pd.DataFrame, user: User) -> Dict[str, int]:
        """Import subtasks data to database."""

        imported = 0
        updated = 0
        skipped = 0

        # Get user_id as integer to avoid relationship loading
        user_id = user.id if user else None
        if not user_id:
            raise ValueError("User ID is required for import")

        # Get application ID mappings
        app_result = await db.execute(select(Application.id, Application.l2_id))
        app_id_map = {l2_id: app_id for app_id, l2_id in app_result.all()}
        print(f"DEBUG: Found {len(app_id_map)} existing applications for mapping")

        # Store original autoflush state and disable it
        original_autoflush = db.autoflush
        db.autoflush = False

        # Collections for batch operations
        new_applications = []  # Applications to create
        new_subtasks = []      # Subtasks to create
        updated_subtasks = []  # Existing subtasks to update

        try:
            for index, row in df.iterrows():
                try:
                    app_l2_id = row.get('l2_id')  # Now looking for l2_id field in Excel

                    # Skip rows with empty L2 ID, including NaN values
                    if pd.isna(app_l2_id) or app_l2_id == '' or app_l2_id is None:
                        skipped += 1
                        if skipped <= 5:  # Log first 5 skipped rows
                            print(f"DEBUG: Skipping row {index + 2} - empty L2 ID")
                        continue

                    # Convert app_l2_id to string (since Application.l2_id is a String field)
                    if isinstance(app_l2_id, (int, float)):
                        import math
                        if math.isnan(app_l2_id):
                            skipped += 1
                            print(f"DEBUG: Skipping row {index + 2} - NaN L2 ID")
                            continue
                        app_l2_id = str(int(app_l2_id))  # Convert numeric to string
                    elif isinstance(app_l2_id, str):
                        app_l2_id = app_l2_id.strip()
                        if not app_l2_id:
                            skipped += 1
                            print(f"DEBUG: Skipping row {index + 2} - empty L2 ID after stripping")
                            continue
                    else:
                        skipped += 1
                        print(f"DEBUG: Skipping row {index + 2} - unexpected L2 ID type: {type(app_l2_id)}")
                        continue

                    application_id = app_id_map.get(app_l2_id)

                    # 如果应用不存在，自动创建placeholder应用
                    if not application_id:
                        print(f"DEBUG: Application with L2 ID '{app_l2_id}' not found, will create placeholder application")

                        # Create placeholder application data
                        new_app_data = {
                            'l2_id': str(app_l2_id),  # Ensure l2_id is a string
                            'app_name': '未命名应用',  # Default name when creating placeholder
                            'current_status': '待启动',
                            'overall_transformation_target': 'AK',
                            'current_transformation_phase': '待启动',
                            'dev_team': '待分配',
                            'dev_owner': '待分配',
                            'ak_supervision_acceptance_year': 2024,  # Add required field with default value
                            'created_by': user_id,  # Use integer directly
                            'updated_by': user_id   # Use integer directly
                        }

                        new_application = Application(**new_app_data)
                        new_applications.append(new_application)
                        # Use a temporary negative ID for mapping (will be replaced after flush)
                        temp_app_id = -(len(new_applications))  # Use negative IDs for temp mapping
                        app_id_map[app_l2_id] = temp_app_id
                        application_id = temp_app_id

                    # Only check existing subtasks if application already exists
                    if application_id > 0:  # Positive IDs are existing applications
                        # Check if subtask exists (based on l2_id and sub_target)
                        stmt = select(SubTask).where(
                            and_(
                                SubTask.l2_id == application_id,
                                SubTask.sub_target == row.get('sub_target')
                            )
                        )
                        result = await db.execute(stmt)
                        existing_subtask = result.scalar_one_or_none()

                        if existing_subtask:
                            # Update existing subtask
                            subtask_model_fields = {
                                'sub_target', 'version_name', 'task_status',
                                'progress_percentage', 'is_blocked', 'block_reason',
                                'planned_requirement_date', 'planned_release_date', 'planned_tech_online_date', 'planned_biz_online_date',
                                'actual_requirement_date', 'actual_release_date', 'actual_tech_online_date', 'actual_biz_online_date',
                                'notes', 'resource_applied', 'ops_testing_status', 'launch_check_status'
                                # Removed ops_requirement_submitted from here since it's a DateTime field that needs special handling
                            }

                            for field, value in row.items():
                                if field in subtask_model_fields and value is not None and value != '' and hasattr(existing_subtask, field):
                                    # Check for NaN in numeric fields
                                    if field == 'progress_percentage' and isinstance(value, float):
                                        import math
                                        if math.isnan(value):
                                            value = 0
                                    setattr(existing_subtask, field, value)

                            # Handle ops_requirement_submitted separately - it's a DateTime field
                            if 'ops_requirement_submitted' in row and hasattr(existing_subtask, 'ops_requirement_submitted'):
                                ops_req_value = row['ops_requirement_submitted']
                                if ops_req_value and not isinstance(ops_req_value, str):
                                    # Only set if it's a valid date/datetime value, not a string like '已完成'
                                    setattr(existing_subtask, 'ops_requirement_submitted', ops_req_value)

                            existing_subtask.updated_by = user_id
                            updated_subtasks.append(existing_subtask)
                            updated += 1
                            continue  # Skip to next row

                    # Create new subtask
                    subtask_data = {}
                    subtask_model_fields = {
                        'sub_target', 'version_name', 'task_status',
                        'progress_percentage', 'is_blocked', 'block_reason',
                        'planned_requirement_date', 'planned_release_date', 'planned_tech_online_date', 'planned_biz_online_date',
                        'actual_requirement_date', 'actual_release_date', 'actual_tech_online_date', 'actual_biz_online_date',
                        'notes', 'resource_applied', 'ops_testing_status', 'launch_check_status'
                        # Removed ops_requirement_submitted from here since it's a DateTime field that needs special handling
                    }

                    for k, v in row.items():
                        if k in subtask_model_fields and v is not None and v != '':
                            subtask_data[k] = v

                    # Handle ops_requirement_submitted separately - it's a DateTime field
                    if 'ops_requirement_submitted' in row:
                        ops_req_value = row['ops_requirement_submitted']
                        if ops_req_value and not isinstance(ops_req_value, str):
                            # Only set if it's a valid date/datetime value, not a string like '已完成'
                            subtask_data['ops_requirement_submitted'] = ops_req_value
                        # If it's a string like '已完成', we ignore it as it's likely mismatched column data

                    # Store the app_l2_id for later resolution
                    subtask_data['_app_l2_id'] = app_l2_id  # Temporary field to track which app this belongs to
                    subtask_data['created_by'] = user_id
                    subtask_data['updated_by'] = user_id

                    # Set defaults
                    if 'sub_target' not in subtask_data or not subtask_data['sub_target']:
                        subtask_data['sub_target'] = 'AK'
                    if 'task_status' not in subtask_data or not subtask_data['task_status']:
                        subtask_data['task_status'] = '待启动'
                    if 'progress_percentage' not in subtask_data:
                        subtask_data['progress_percentage'] = 0
                    else:
                        # Check for NaN in progress_percentage
                        import math
                        if isinstance(subtask_data['progress_percentage'], float) and math.isnan(subtask_data['progress_percentage']):
                            subtask_data['progress_percentage'] = 0
                    if 'is_blocked' not in subtask_data:
                        subtask_data['is_blocked'] = False
                    if 'resource_applied' not in subtask_data:
                        subtask_data['resource_applied'] = False

                    new_subtasks.append(subtask_data)  # Store as dict for now
                    imported += 1

                except Exception as e:
                    print(f"Error processing subtask row {index + 2}: {e}")
                    import traceback
                    traceback.print_exc()
                    skipped += 1
                    continue

            # Now add all new applications in batch
            if new_applications:
                print(f"DEBUG: Adding {len(new_applications)} new placeholder applications...")
                for app in new_applications:
                    db.add(app)

                # Flush to get real IDs for the applications
                await db.flush()

                # Update the mapping with real IDs
                for app in new_applications:
                    # Find the temporary mapping and update it
                    for l2_id, temp_id in list(app_id_map.items()):
                        if temp_id < 0:  # It's a temporary ID
                            # Check if this is the right app by matching l2_id
                            if app.l2_id == l2_id:
                                app_id_map[l2_id] = app.id
                                print(f"DEBUG: Mapped L2 ID '{l2_id}' to real application ID {app.id}")
                                break

            # Now create SubTask objects with correct application IDs
            if new_subtasks:
                print(f"DEBUG: Creating {len(new_subtasks)} new subtasks...")
                for subtask_data in new_subtasks:
                    # Get the real application ID
                    app_l2_id = subtask_data.pop('_app_l2_id', None)
                    if app_l2_id:
                        real_app_id = app_id_map.get(app_l2_id)
                        if real_app_id and real_app_id > 0:
                            subtask_data['l2_id'] = real_app_id
                        else:
                            print(f"WARNING: Could not find application ID for L2 ID '{app_l2_id}'")
                            continue

                    # Create and add the subtask
                    new_subtask = SubTask(**subtask_data)
                    db.add(new_subtask)

            # Commit all changes at once
            await db.commit()
            print(f"DEBUG: Successfully committed all changes")

        except Exception as e:
            # Restore autoflush and rollback on error
            db.autoflush = original_autoflush
            await db.rollback()
            print(f"ERROR: Failed to import subtasks: {e}")
            import traceback
            traceback.print_exc()
            raise

        finally:
            # Always restore autoflush state
            db.autoflush = original_autoflush

        # Log final statistics
        print(f"DEBUG: Import completed - Imported: {imported}, Updated: {updated}, Skipped: {skipped}")
        if skipped > 0:
            print(f"DEBUG: {skipped} rows were skipped")

        return {
            'imported': imported,
            'updated': updated,
            'skipped': skipped
        }

    async def _get_applications_for_export(
        self,
        db: AsyncSession,
        application_ids: Optional[List[int]] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Application]:
        """Get applications for export."""

        query = select(Application)

        if application_ids:
            query = query.where(Application.id.in_(application_ids))

        if filters:
            if filters.get('supervision_year'):
                query = query.where(Application.supervision_year == filters['supervision_year'])
            if filters.get('responsible_team'):
                query = query.where(Application.responsible_team == filters['responsible_team'])
            if filters.get('overall_status'):
                query = query.where(Application.overall_status == filters['overall_status'])

        result = await db.execute(query.order_by(Application.l2_id))
        return result.scalars().all()

    async def _get_subtasks_for_export(
            self,
            db: AsyncSession,
            application_id: Optional[int] = None,
            subtask_ids: Optional[List[int]] = None,
            filters: Optional[Dict[str, Any]] = None
    ) -> List[SubTask]:
        """Get subtasks for export."""

        # 修正：SubTask使用l2_id关联Application，不是application_id
        query = select(SubTask).join(Application, SubTask.l2_id == Application.id)

        if application_id:
            query = query.where(SubTask.l2_id == application_id)

        if subtask_ids:
            query = query.where(SubTask.id.in_(subtask_ids))

        if filters:
            if filters.get('task_status'):
                query = query.where(SubTask.task_status == filters['task_status'])
            if filters.get('sub_target'):
                query = query.where(SubTask.sub_target == filters['sub_target'])
            if filters.get('is_blocked') is not None:
                query = query.where(SubTask.is_blocked == filters['is_blocked'])

        # 修正排序字段：使用version_name而不是module_name
        result = await db.execute(query.order_by(Application.l2_id, SubTask.version_name))
        return result.scalars().all()

    def _write_headers(self, worksheet, headers: List[str], style: str = "standard"):
        """Write headers to worksheet."""

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)

            # Apply header styling
            if style == "standard":
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _format_cell_value(self, value: Any, field_name: str) -> Any:
        """Format cell value for Excel export."""

        if value is None:
            return ''

        if field_name in self.config.DATE_FIELDS:
            if isinstance(value, date):
                return value.strftime('%Y-%m-%d')
        elif field_name in self.config.BOOLEAN_FIELDS:
            if isinstance(value, bool):
                return '是' if value else '否'

        return value

    def _apply_worksheet_styling(self, worksheet, max_row: int, max_col: int, style: str = "standard"):
        """Apply styling to worksheet."""

        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in range(1, max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length

            # Set column width
            adjusted_width = min(max(max_length + 2, 8), 50)  # Min 8, max 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border

    def _create_applications_template(self, workbook: Workbook, include_sample: bool = False):
        """Create applications import template."""

        worksheet = workbook.active
        worksheet.title = "应用导入模板"

        # Headers
        headers = list(self.config.APPLICATION_FIELDS.keys())
        self._write_headers(worksheet, headers, "standard")

        if include_sample:
            # Sample data
            sample_data = [
                'L2_APP_001', '支付系统', 2024, 'AK', '开发阶段', '研发进行中',
                '核心技术团队', '张三', 60, '2024-01-15', '2024-03-01',
                '2024-03-15', '2024-04-01', '', '', '', '', '这是示例应用'
            ]

            for col, value in enumerate(sample_data, 1):
                worksheet.cell(row=2, column=col, value=value)

        # Apply styling
        self._apply_worksheet_styling(worksheet, 2 if include_sample else 1, len(headers))

    def _create_subtasks_template(self, workbook: Workbook, include_sample: bool = False):
        """Create subtasks import template."""

        worksheet = workbook.active
        worksheet.title = "子任务导入模板"

        # Headers - 使用实际存在的字段
        headers = [
            'L2ID', 'L2应用', '子目标', '版本名', '改造状态',
            '进度百分比', '是否阻塞', '阻塞原因',
            '【计划】需求完成时间', '【计划】发版时间',
            '【计划】技术上线时间', '【计划】业务上线时间',
            '【实际】需求到达时间', '【实际】发版时间',
            '【实际】技术上线时间', '【实际】业务上线时间',
            '资源是否申请', '运营需求提交', '运营测试', '上线检查', '备注'
        ]
        self._write_headers(worksheet, headers, "standard")

        if include_sample:
            # 更新示例数据
            sample_data = [
                'CI001072775',  # L2_ID
                '云闪付数据中心分流应用',  # app_name
                'AK',  # sub_target
                'v1.0',  # version_name
                '研发进行中',  # task_status
                80,  # progress_percentage
                '否',  # is_blocked
                '',  # block_reason
                '2024-01-15',  # planned dates...
                '2024-02-15',
                '2024-02-28',
                '2024-03-15',
                '',  # actual dates...
                '',
                '',
                '',
                '否',  # resource_applied
                '',  # ops_requirement_submitted
                '',  # ops_testing_status
                '',  # launch_check_status
                '示例备注'  # notes
            ]

            for col, value in enumerate(sample_data, 1):
                worksheet.cell(row=2, column=col, value=value)

        # Apply styling
        self._apply_worksheet_styling(worksheet, 2 if include_sample else 1, len(headers))

    def _create_combined_template(self, workbook: Workbook, include_sample: bool = False):
        """Create combined template with both applications and subtasks."""

        # Create applications sheet
        apps_sheet = workbook.active
        apps_sheet.title = "应用列表"

        headers = list(self.config.APPLICATION_FIELDS.keys())
        self._write_headers(apps_sheet, headers, "standard")

        if include_sample:
            sample_data = [
                'L2_APP_001', '支付系统', 2024, 'AK', '开发阶段', '研发进行中',
                '核心技术团队', '张三', 60, '2024-01-15', '2024-03-01',
                '2024-03-15', '2024-04-01', '', '', '', '', '这是示例应用'
            ]

            for col, value in enumerate(sample_data, 1):
                apps_sheet.cell(row=2, column=col, value=value)

        self._apply_worksheet_styling(apps_sheet, 2 if include_sample else 1, len(headers))

        # Create subtasks sheet
        subtasks_sheet = workbook.create_sheet("子任务列表")

        headers = list(self.config.SUBTASK_FIELDS.keys())
        self._write_headers(subtasks_sheet, headers, "standard")

        if include_sample:
            sample_data = [
                'L2_APP_001', '用户认证模块', 'AK', 'v1.0', '研发进行中',
                80, '否', '', '2024-01-15', '2024-02-15', '2024-02-28',
                '2024-03-15', '', '', '', '', 40, '认证功能开发'
            ]

            for col, value in enumerate(sample_data, 1):
                subtasks_sheet.cell(row=2, column=col, value=value)

        self._apply_worksheet_styling(subtasks_sheet, 2 if include_sample else 1, len(headers))