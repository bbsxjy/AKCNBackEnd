"""
Audit service layer for tracking and managing audit logs
"""

from typing import List, Dict, Any, Optional, Tuple, Union
from datetime import datetime, date
from sqlalchemy import select, func, and_, or_, desc, asc, text, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models.audit_log import AuditLog, AuditOperation
from app.models.user import User
from app.models.application import Application
from app.models.subtask import SubTask
from app.core.exceptions import NotFoundError, ValidationError


class AuditService:
    """Audit service for tracking and managing audit logs."""

    def __init__(self):
        self.model = AuditLog

    async def create_audit_log(
        self,
        db: AsyncSession,
        table_name: str,
        record_id: int,
        operation: AuditOperation,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None,
        user_id: Optional[int] = None,
        request_id: Optional[str] = None,
        user_ip: Optional[str] = None,
        user_agent: Optional[str] = None,
        reason: Optional[str] = None,
        extra_data: Optional[Dict[str, Any]] = None
    ) -> AuditLog:
        """Create a new audit log entry."""

        # Calculate changed fields for UPDATE operations
        changed_fields = None
        if operation == AuditOperation.UPDATE and old_values and new_values:
            changed_fields = []
            for key in new_values:
                if key in old_values and old_values[key] != new_values[key]:
                    changed_fields.append(key)

        audit_log = AuditLog(
            table_name=table_name,
            record_id=record_id,
            operation=operation.value,
            old_values=old_values,
            new_values=new_values,
            changed_fields=changed_fields,
            request_id=request_id,
            user_ip=user_ip,
            user_agent=user_agent,
            reason=reason,
            extra_data=extra_data,
            user_id=user_id
        )

        db.add(audit_log)
        await db.commit()
        await db.refresh(audit_log)
        return audit_log

    async def get_audit_log(self, db: AsyncSession, audit_log_id: int) -> Optional[AuditLog]:
        """Get audit log by ID."""
        result = await db.execute(
            select(AuditLog)
            .options(selectinload(AuditLog.user))
            .where(AuditLog.id == audit_log_id)
        )
        return result.scalar_one_or_none()

    async def list_audit_logs(
        self,
        db: AsyncSession,
        skip: int = 0,
        limit: int = 100,
        table_name: Optional[str] = None,
        record_id: Optional[int] = None,
        operation: Optional[AuditOperation] = None,
        user_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        search: Optional[str] = None
    ) -> Tuple[List[AuditLog], int]:
        """List audit logs with filtering and pagination."""

        # Build base query
        query = select(AuditLog).options(selectinload(AuditLog.user))
        count_query = select(func.count(AuditLog.id))

        conditions = []

        # Apply filters
        if table_name:
            conditions.append(AuditLog.table_name == table_name)

        if record_id:
            conditions.append(AuditLog.record_id == record_id)

        if operation:
            conditions.append(AuditLog.operation == operation.value)

        if user_id:
            conditions.append(AuditLog.user_id == user_id)

        if start_date:
            conditions.append(AuditLog.created_at >= start_date)

        if end_date:
            # End of day for end_date
            end_datetime = datetime.combine(end_date, datetime.max.time())
            conditions.append(AuditLog.created_at <= end_datetime)

        if search:
            # Search in reason, extra_data, or user agent
            search_conditions = [
                AuditLog.reason.ilike(f"%{search}%"),
                AuditLog.user_agent.ilike(f"%{search}%"),
                AuditLog.request_id.ilike(f"%{search}%")
            ]
            conditions.append(or_(*search_conditions))

        # Apply all conditions
        if conditions:
            query = query.where(and_(*conditions))
            count_query = count_query.where(and_(*conditions))

        # Get total count
        total_result = await db.execute(count_query)
        total = total_result.scalar()

        # Apply sorting and pagination
        query = query.order_by(desc(AuditLog.created_at))
        query = query.offset(skip).limit(limit)

        # Execute query
        result = await db.execute(query)
        audit_logs = result.scalars().all()

        return audit_logs, total

    async def get_record_history(
        self,
        db: AsyncSession,
        table_name: str,
        record_id: int
    ) -> List[AuditLog]:
        """Get complete history for a specific record."""
        result = await db.execute(
            select(AuditLog)
            .options(selectinload(AuditLog.user))
            .where(
                and_(
                    AuditLog.table_name == table_name,
                    AuditLog.record_id == record_id
                )
            )
            .order_by(desc(AuditLog.created_at))
        )
        return result.scalars().all()

    async def get_user_activity(
        self,
        db: AsyncSession,
        user_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: int = 100
    ) -> List[AuditLog]:
        """Get audit logs for a specific user."""
        conditions = [AuditLog.user_id == user_id]

        if start_date:
            conditions.append(AuditLog.created_at >= start_date)

        if end_date:
            end_datetime = datetime.combine(end_date, datetime.max.time())
            conditions.append(AuditLog.created_at <= end_datetime)

        result = await db.execute(
            select(AuditLog)
            .options(selectinload(AuditLog.user))
            .where(and_(*conditions))
            .order_by(desc(AuditLog.created_at))
            .limit(limit)
        )
        return result.scalars().all()

    async def get_audit_statistics(
        self,
        db: AsyncSession,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """Get audit log statistics."""

        conditions = []
        if start_date:
            conditions.append(AuditLog.created_at >= start_date)
        if end_date:
            end_datetime = datetime.combine(end_date, datetime.max.time())
            conditions.append(AuditLog.created_at <= end_datetime)

        base_filter = and_(*conditions) if conditions else True

        # Total audit logs
        total_result = await db.execute(
            select(func.count(AuditLog.id)).where(base_filter)
        )
        total_logs = total_result.scalar()

        # Operations breakdown
        operations_result = await db.execute(
            select(AuditLog.operation, func.count(AuditLog.id))
            .where(base_filter)
            .group_by(AuditLog.operation)
        )
        by_operation = {op: count for op, count in operations_result.all()}

        # Tables breakdown
        tables_result = await db.execute(
            select(AuditLog.table_name, func.count(AuditLog.id))
            .where(base_filter)
            .group_by(AuditLog.table_name)
        )
        by_table = {table: count for table, count in tables_result.all()}

        # Top users
        users_result = await db.execute(
            select(AuditLog.user_id, func.count(AuditLog.id))
            .where(and_(base_filter, AuditLog.user_id.isnot(None)))
            .group_by(AuditLog.user_id)
            .order_by(desc(func.count(AuditLog.id)))
            .limit(10)
        )
        top_users = [{"user_id": user_id, "count": count} for user_id, count in users_result.all()]

        # Activity by hour (for recent activity patterns)
        if not start_date:
            # Default to last 7 days for hourly analysis
            from datetime import timedelta
            start_date = date.today() - timedelta(days=7)

        hourly_result = await db.execute(
            text("""
                SELECT EXTRACT(HOUR FROM created_at) as hour, COUNT(*) as count
                FROM audit_logs
                WHERE created_at >= :start_date
                GROUP BY EXTRACT(HOUR FROM created_at)
                ORDER BY hour
            """),
            {"start_date": start_date}
        )
        by_hour = {int(hour): count for hour, count in hourly_result.all()}

        return {
            "total_logs": total_logs,
            "by_operation": by_operation,
            "by_table": by_table,
            "top_users": top_users,
            "activity_by_hour": by_hour,
            "period_start": start_date.isoformat() if start_date else None,
            "period_end": end_date.isoformat() if end_date else None
        }

    async def get_data_changes_summary(
        self,
        db: AsyncSession,
        table_name: str,
        record_id: int
    ) -> Dict[str, Any]:
        """Get summary of all changes made to a specific record."""

        # Get all audit logs for the record
        history = await self.get_record_history(db, table_name, record_id)

        if not history:
            return {
                "table_name": table_name,
                "record_id": record_id,
                "total_changes": 0,
                "created_at": None,
                "last_modified_at": None,
                "created_by": None,
                "last_modified_by": None,
                "change_summary": {}
            }

        # Find creation and latest modification
        creation_log = None
        latest_log = history[0]  # Already sorted by created_at desc

        for log in reversed(history):  # Check from oldest
            if log.operation == AuditOperation.INSERT:
                creation_log = log
                break

        # Count changes per field
        field_changes = {}
        total_changes = 0

        for log in history:
            if log.operation == AuditOperation.UPDATE and log.changed_fields:
                total_changes += 1
                for field in log.changed_fields:
                    if field not in field_changes:
                        field_changes[field] = 0
                    field_changes[field] += 1

        return {
            "table_name": table_name,
            "record_id": record_id,
            "total_changes": total_changes,
            "total_operations": len(history),
            "created_at": creation_log.created_at.isoformat() if creation_log else None,
            "last_modified_at": latest_log.created_at.isoformat(),
            "created_by": creation_log.user_id if creation_log else None,
            "last_modified_by": latest_log.user_id,
            "operations_breakdown": {
                "INSERT": len([l for l in history if l.operation == AuditOperation.INSERT]),
                "UPDATE": len([l for l in history if l.operation == AuditOperation.UPDATE]),
                "DELETE": len([l for l in history if l.operation == AuditOperation.DELETE])
            },
            "field_changes": field_changes,
            "most_changed_fields": sorted(field_changes.items(), key=lambda x: x[1], reverse=True)[:5]
        }

    async def cleanup_old_logs(
        self,
        db: AsyncSession,
        days_to_keep: int = 365
    ) -> int:
        """Clean up old audit logs beyond retention period."""
        from datetime import timedelta

        cutoff_date = date.today() - timedelta(days=days_to_keep)

        # Count logs to be deleted
        count_result = await db.execute(
            select(func.count(AuditLog.id))
            .where(AuditLog.created_at < cutoff_date)
        )
        count_to_delete = count_result.scalar()

        if count_to_delete > 0:
            # Delete old logs
            await db.execute(
                text("DELETE FROM audit_logs WHERE created_at < :cutoff_date"),
                {"cutoff_date": cutoff_date}
            )
            await db.commit()

        return count_to_delete

    async def export_audit_trail(
        self,
        db: AsyncSession,
        table_name: Optional[str] = None,
        record_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        format: str = "json"
    ) -> List[Dict[str, Any]]:
        """Export audit trail data for compliance purposes."""

        # Build query
        query = select(AuditLog).options(selectinload(AuditLog.user))
        conditions = []

        if table_name:
            conditions.append(AuditLog.table_name == table_name)
        if record_id:
            conditions.append(AuditLog.record_id == record_id)
        if start_date:
            conditions.append(AuditLog.created_at >= start_date)
        if end_date:
            end_datetime = datetime.combine(end_date, datetime.max.time())
            conditions.append(AuditLog.created_at <= end_datetime)

        if conditions:
            query = query.where(and_(*conditions))

        query = query.order_by(asc(AuditLog.created_at))

        result = await db.execute(query)
        audit_logs = result.scalars().all()

        # Format for export
        export_data = []
        for log in audit_logs:
            export_record = {
                "id": log.id,
                "timestamp": log.created_at.isoformat(),
                "table_name": log.table_name,
                "record_id": log.record_id,
                "operation": log.operation,
                "user_id": log.user_id,
                "username": log.user.username if log.user else None,
                "user_full_name": log.user.full_name if log.user else None,
                "changed_fields": log.changed_fields,
                "field_changes": log.get_field_changes(),
                "old_values": log.old_values,
                "new_values": log.new_values,
                "request_id": log.request_id,
                "user_ip": log.user_ip,
                "user_agent": log.user_agent,
                "reason": log.reason,
                "extra_data": log.extra_data
            }
            export_data.append(export_record)

        return export_data

    async def export_audit_trail_to_csv(
        self,
        db: AsyncSession,
        table_name: Optional[str] = None,
        record_id: Optional[int] = None,
        user_id: Optional[int] = None,
        operation: Optional[AuditOperation] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """Export audit trail to CSV format."""

        # Get audit data
        export_data = await self.export_audit_trail(
            db=db,
            table_name=table_name,
            record_id=record_id,
            start_date=start_date,
            end_date=end_date
        )

        # Create CSV in memory
        output = io.StringIO()

        if export_data:
            # Define CSV columns
            fieldnames = [
                'id', 'timestamp', 'table_name', 'record_id', 'operation',
                'user_id', 'username', 'user_full_name', 'changed_fields',
                'request_id', 'user_ip', 'reason'
            ]

            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()

            for log in export_data:
                writer.writerow({
                    'id': log['id'],
                    'timestamp': log['timestamp'],
                    'table_name': log['table_name'],
                    'record_id': log['record_id'],
                    'operation': log['operation'],
                    'user_id': log['user_id'],
                    'username': log['username'],
                    'user_full_name': log['user_full_name'],
                    'changed_fields': json.dumps(log['changed_fields']) if log['changed_fields'] else '',
                    'request_id': log['request_id'],
                    'user_ip': log['user_ip'],
                    'reason': log['reason']
                })

        # Convert to bytes
        output.seek(0)
        return output.getvalue().encode('utf-8-sig')  # UTF-8 with BOM for Excel compatibility

    async def export_audit_trail_to_excel(
        self,
        db: AsyncSession,
        table_name: Optional[str] = None,
        record_id: Optional[int] = None,
        user_id: Optional[int] = None,
        operation: Optional[AuditOperation] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """Export audit trail to Excel format."""

        # Get audit data
        export_data = await self.export_audit_trail(
            db=db,
            table_name=table_name,
            record_id=record_id,
            start_date=start_date,
            end_date=end_date
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Trail"

        # Define headers
        headers = [
            'ID', 'Timestamp', 'Table', 'Record ID', 'Operation',
            'User ID', 'Username', 'Full Name', 'Changed Fields',
            'Old Values', 'New Values', 'Request ID', 'IP Address', 'Reason'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_idx, log in enumerate(export_data, 2):
            ws.cell(row=row_idx, column=1, value=log['id'])
            ws.cell(row=row_idx, column=2, value=log['timestamp'])
            ws.cell(row=row_idx, column=3, value=log['table_name'])
            ws.cell(row=row_idx, column=4, value=log['record_id'])
            ws.cell(row=row_idx, column=5, value=log['operation'])
            ws.cell(row=row_idx, column=6, value=log['user_id'])
            ws.cell(row=row_idx, column=7, value=log['username'])
            ws.cell(row=row_idx, column=8, value=log['user_full_name'])
            ws.cell(row=row_idx, column=9, value=json.dumps(log['changed_fields']) if log['changed_fields'] else '')
            ws.cell(row=row_idx, column=10, value=json.dumps(log['old_values']) if log['old_values'] else '')
            ws.cell(row=row_idx, column=11, value=json.dumps(log['new_values']) if log['new_values'] else '')
            ws.cell(row=row_idx, column=12, value=log['request_id'])
            ws.cell(row=row_idx, column=13, value=log['user_ip'])
            ws.cell(row=row_idx, column=14, value=log['reason'])

        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # Add filters
        ws.auto_filter.ref = ws.dimensions

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Export Summary"])
        ws_summary.append(["Total Records", len(export_data)])
        ws_summary.append(["Export Date", datetime.utcnow().isoformat()])

        if start_date:
            ws_summary.append(["Start Date", start_date.isoformat()])
        if end_date:
            ws_summary.append(["End Date", end_date.isoformat()])
        if table_name:
            ws_summary.append(["Table Filter", table_name])
        if record_id:
            ws_summary.append(["Record ID Filter", record_id])

        # Operation breakdown
        if export_data:
            operation_counts = {}
            for log in export_data:
                op = log['operation']
                operation_counts[op] = operation_counts.get(op, 0) + 1

            ws_summary.append([])
            ws_summary.append(["Operation Breakdown"])
            for op, count in operation_counts.items():
                ws_summary.append([op, count])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def get_compliance_report(
        self,
        db: AsyncSession,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate compliance report for audit trail."""

        conditions = [
            AuditLog.created_at >= start_date,
            AuditLog.created_at <= datetime.combine(end_date, datetime.max.time())
        ]
        base_filter = and_(*conditions)

        # Basic statistics
        stats = await self.get_audit_statistics(db, start_date, end_date)

        # Data integrity checks
        integrity_checks = {
            "logs_without_user": 0,
            "logs_with_changes_but_no_fields": 0,
            "suspicious_bulk_operations": 0
        }

        # Count logs without user (system operations)
        no_user_result = await db.execute(
            select(func.count(AuditLog.id))
            .where(and_(base_filter, AuditLog.user_id.is_(None)))
        )
        integrity_checks["logs_without_user"] = no_user_result.scalar()

        # Count UPDATE operations with values but no changed_fields
        inconsistent_result = await db.execute(
            select(func.count(AuditLog.id))
            .where(and_(
                base_filter,
                AuditLog.operation == AuditOperation.UPDATE,
                AuditLog.new_values.isnot(None),
                AuditLog.changed_fields.is_(None)
            ))
        )
        integrity_checks["logs_with_changes_but_no_fields"] = inconsistent_result.scalar()

        # Detect potential bulk operations (same user, same minute, multiple records)
        bulk_ops_result = await db.execute(
            text("""
                SELECT user_id, DATE_TRUNC('minute', created_at) as minute, COUNT(*) as count
                FROM audit_logs
                WHERE created_at BETWEEN :start_date AND :end_date
                AND user_id IS NOT NULL
                GROUP BY user_id, DATE_TRUNC('minute', created_at)
                HAVING COUNT(*) > 10
                ORDER BY count DESC
                LIMIT 10
            """),
            {"start_date": start_date, "end_date": end_date}
        )
        bulk_operations = [
            {"user_id": user_id, "minute": minute.isoformat(), "count": count}
            for user_id, minute, count in bulk_ops_result.all()
        ]
        integrity_checks["suspicious_bulk_operations"] = len(bulk_operations)

        return {
            "report_period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat()
            },
            "statistics": stats,
            "integrity_checks": integrity_checks,
            "bulk_operations": bulk_operations,
            "coverage": {
                "tables_with_audit": len(stats["by_table"]),
                "users_with_activity": len(stats["top_users"])
            },
            "generated_at": datetime.utcnow().isoformat()
        }

    async def rollback_change(
        self,
        db: AsyncSession,
        audit_log_id: int,
        user_id: int,
        reason: Optional[str] = None,
        request_id: Optional[str] = None,
        user_ip: Optional[str] = None,
        user_agent: Optional[str] = None
    ) -> Dict[str, Any]:
        """Rollback a change based on audit log entry."""

        # Get the audit log entry
        audit_log = await self.get_audit_log(db, audit_log_id)
        if not audit_log:
            raise NotFoundError(f"Audit log with ID {audit_log_id} not found")

        # Validate that the operation can be rolled back
        if audit_log.operation == AuditOperation.INSERT:
            # For INSERT, we need to DELETE the record
            return await self._rollback_insert(db, audit_log, user_id, reason, request_id, user_ip, user_agent)
        elif audit_log.operation == AuditOperation.UPDATE:
            # For UPDATE, we need to restore old values
            return await self._rollback_update(db, audit_log, user_id, reason, request_id, user_ip, user_agent)
        elif audit_log.operation == AuditOperation.DELETE:
            # For DELETE, we need to INSERT the record back
            return await self._rollback_delete(db, audit_log, user_id, reason, request_id, user_ip, user_agent)
        else:
            raise ValidationError(f"Unknown operation type: {audit_log.operation}")

    async def _rollback_insert(
        self,
        db: AsyncSession,
        audit_log: AuditLog,
        user_id: int,
        reason: Optional[str],
        request_id: Optional[str],
        user_ip: Optional[str],
        user_agent: Optional[str]
    ) -> Dict[str, Any]:
        """Rollback an INSERT operation by deleting the record."""

        # Get the model class for the table
        model_class = self._get_model_class(audit_log.table_name)
        if not model_class:
            raise ValidationError(f"Unknown table: {audit_log.table_name}")

        # Find and delete the record
        result = await db.execute(
            select(model_class).where(model_class.id == audit_log.record_id)
        )
        record = result.scalar_one_or_none()

        if not record:
            raise NotFoundError(f"Record {audit_log.record_id} not found in table {audit_log.table_name}")

        # Store current values before deletion
        current_values = self._serialize_record(record)

        # Delete the record
        await db.delete(record)

        # Create rollback audit log
        rollback_audit = await self.create_audit_log(
            db=db,
            table_name=audit_log.table_name,
            record_id=audit_log.record_id,
            operation=AuditOperation.DELETE,
            old_values=current_values,
            new_values=None,
            user_id=user_id,
            request_id=request_id,
            user_ip=user_ip,
            user_agent=user_agent,
            reason=f"Rollback of INSERT (audit_id={audit_log.id}): {reason}" if reason else f"Rollback of INSERT (audit_id={audit_log.id})",
            extra_data={"rollback_of_audit_id": audit_log.id}
        )

        await db.commit()

        return {
            "status": "success",
            "rollback_audit_id": rollback_audit.id,
            "affected_record": {
                "table": audit_log.table_name,
                "id": audit_log.record_id,
                "operation": "DELETE",
                "restored_values": None
            }
        }

    async def _rollback_update(
        self,
        db: AsyncSession,
        audit_log: AuditLog,
        user_id: int,
        reason: Optional[str],
        request_id: Optional[str],
        user_ip: Optional[str],
        user_agent: Optional[str]
    ) -> Dict[str, Any]:
        """Rollback an UPDATE operation by restoring old values."""

        if not audit_log.old_values:
            raise ValidationError("Cannot rollback UPDATE: no old values stored")

        # Get the model class for the table
        model_class = self._get_model_class(audit_log.table_name)
        if not model_class:
            raise ValidationError(f"Unknown table: {audit_log.table_name}")

        # Find the record
        result = await db.execute(
            select(model_class).where(model_class.id == audit_log.record_id)
        )
        record = result.scalar_one_or_none()

        if not record:
            raise NotFoundError(f"Record {audit_log.record_id} not found in table {audit_log.table_name}")

        # Store current values before update
        current_values = self._serialize_record(record)

        # Restore old values
        for field, value in audit_log.old_values.items():
            if hasattr(record, field):
                # Handle datetime fields
                if field in ['created_at', 'updated_at', 'start_date', 'end_date', 'planned_online_date', 'actual_online_date']:
                    if value:
                        value = datetime.fromisoformat(value.replace('Z', '+00:00')) if isinstance(value, str) else value
                setattr(record, field, value)

        # Create rollback audit log
        rollback_audit = await self.create_audit_log(
            db=db,
            table_name=audit_log.table_name,
            record_id=audit_log.record_id,
            operation=AuditOperation.UPDATE,
            old_values=current_values,
            new_values=audit_log.old_values,
            user_id=user_id,
            request_id=request_id,
            user_ip=user_ip,
            user_agent=user_agent,
            reason=f"Rollback of UPDATE (audit_id={audit_log.id}): {reason}" if reason else f"Rollback of UPDATE (audit_id={audit_log.id})",
            extra_data={"rollback_of_audit_id": audit_log.id}
        )

        await db.commit()

        return {
            "status": "success",
            "rollback_audit_id": rollback_audit.id,
            "affected_record": {
                "table": audit_log.table_name,
                "id": audit_log.record_id,
                "operation": "UPDATE",
                "restored_values": audit_log.old_values
            }
        }

    async def _rollback_delete(
        self,
        db: AsyncSession,
        audit_log: AuditLog,
        user_id: int,
        reason: Optional[str],
        request_id: Optional[str],
        user_ip: Optional[str],
        user_agent: Optional[str]
    ) -> Dict[str, Any]:
        """Rollback a DELETE operation by re-inserting the record."""

        if not audit_log.old_values:
            raise ValidationError("Cannot rollback DELETE: no old values stored")

        # Get the model class for the table
        model_class = self._get_model_class(audit_log.table_name)
        if not model_class:
            raise ValidationError(f"Unknown table: {audit_log.table_name}")

        # Create new record with old values
        record_data = {}
        for field, value in audit_log.old_values.items():
            if field != 'id' and hasattr(model_class, field):
                # Handle datetime fields
                if field in ['created_at', 'updated_at', 'start_date', 'end_date', 'planned_online_date', 'actual_online_date']:
                    if value:
                        value = datetime.fromisoformat(value.replace('Z', '+00:00')) if isinstance(value, str) else value
                record_data[field] = value

        # Create new record
        new_record = model_class(**record_data)
        new_record.id = audit_log.record_id  # Restore original ID

        db.add(new_record)

        # Create rollback audit log
        rollback_audit = await self.create_audit_log(
            db=db,
            table_name=audit_log.table_name,
            record_id=audit_log.record_id,
            operation=AuditOperation.INSERT,
            old_values=None,
            new_values=audit_log.old_values,
            user_id=user_id,
            request_id=request_id,
            user_ip=user_ip,
            user_agent=user_agent,
            reason=f"Rollback of DELETE (audit_id={audit_log.id}): {reason}" if reason else f"Rollback of DELETE (audit_id={audit_log.id})",
            extra_data={"rollback_of_audit_id": audit_log.id}
        )

        await db.commit()

        return {
            "status": "success",
            "rollback_audit_id": rollback_audit.id,
            "affected_record": {
                "table": audit_log.table_name,
                "id": audit_log.record_id,
                "operation": "INSERT",
                "restored_values": audit_log.old_values
            }
        }

    def _get_model_class(self, table_name: str):
        """Get the SQLAlchemy model class for a table name."""
        # Map table names to model classes
        table_model_map = {
            "applications": Application,
            "sub_tasks": SubTask,
            "users": User,
            "audit_logs": AuditLog
        }
        return table_model_map.get(table_name)

    def _serialize_record(self, record) -> Dict[str, Any]:
        """Serialize a SQLAlchemy model instance to dictionary."""
        result = {}
        for column in record.__table__.columns:
            value = getattr(record, column.name)
            # Handle datetime serialization
            if isinstance(value, datetime):
                value = value.isoformat()
            elif isinstance(value, date):
                value = value.isoformat()
            result[column.name] = value
        return result

    async def log_action(
        self,
        db: AsyncSession,
        user_id: int,
        action: str,
        resource_type: str,
        resource_id: int,
        details: Optional[str] = None,
        old_values: Optional[Dict[str, Any]] = None,
        new_values: Optional[Dict[str, Any]] = None
    ) -> AuditLog:
        """
        Simplified method to log an action.

        Args:
            db: Database session
            user_id: User performing the action
            action: Action type (CREATE, UPDATE, DELETE)
            resource_type: Type of resource (user, application, etc.)
            resource_id: ID of the resource
            details: Optional description
            old_values: Optional old values
            new_values: Optional new values

        Returns:
            Created audit log
        """
        # Map action strings to AuditOperation enum
        operation_map = {
            "CREATE": AuditOperation.INSERT,
            "UPDATE": AuditOperation.UPDATE,
            "DELETE": AuditOperation.DELETE,
            "INSERT": AuditOperation.INSERT
        }

        operation = operation_map.get(action.upper(), AuditOperation.UPDATE)

        return await self.create_audit_log(
            db=db,
            table_name=resource_type,
            record_id=resource_id,
            operation=operation,
            old_values=old_values,
            new_values=new_values,
            user_id=user_id,
            reason=details
        )


# Create singleton instance
audit_service = AuditService()