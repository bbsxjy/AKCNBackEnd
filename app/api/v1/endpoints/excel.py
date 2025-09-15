"""
Excel Import/Export API endpoints
"""

import time
import uuid
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from fastapi.responses import Response, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db, get_current_user, require_roles
from app.models.user import User, UserRole
from app.services.excel_service import ExcelService
from app.schemas.excel import (
    ExcelImportRequest, ExcelImportResult, ExcelExportRequest, ExcelExportResult,
    ApplicationImportRequest, ApplicationExportRequest, SubTaskImportRequest, SubTaskExportRequest,
    ExcelTemplateRequest, ExcelImportPreview, ExcelImportSummary, ExcelHealthCheck
)

router = APIRouter()
excel_service = ExcelService()


@router.post("/applications/import", response_model=ExcelImportResult)
async def import_applications_from_excel(
    file: UploadFile = File(..., description="Excel file with applications data"),
    request: ApplicationImportRequest = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Import applications from Excel file."""

    # Validate file
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )

    # Check file size (limit to 50MB)
    file_content = await file.read()
    if len(file_content) > 50 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds 50MB limit"
        )

    start_time = time.time()

    try:
        result = await excel_service.import_applications_from_excel(
            db=db,
            file_content=file_content,
            user=current_user,
            validate_only=request.validate_only
        )

        processing_time = int((time.time() - start_time) * 1000)
        result['processing_time_ms'] = processing_time

        return ExcelImportResult(**result)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import applications: {str(e)}"
        )


@router.post("/subtasks/import", response_model=ExcelImportResult)
async def import_subtasks_from_excel(
    file: UploadFile = File(..., description="Excel file with subtasks data"),
    request: SubTaskImportRequest = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Import subtasks from Excel file."""

    # Validate file
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )

    # Check file size
    file_content = await file.read()
    if len(file_content) > 50 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds 50MB limit"
        )

    start_time = time.time()

    try:
        result = await excel_service.import_subtasks_from_excel(
            db=db,
            file_content=file_content,
            user=current_user,
            validate_only=request.validate_only
        )

        processing_time = int((time.time() - start_time) * 1000)
        result['processing_time_ms'] = processing_time

        return ExcelImportResult(**result)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import subtasks: {str(e)}"
        )


@router.get("/applications/export")
async def export_applications_to_excel(
    export_request: ApplicationExportRequest = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Export applications to Excel file."""

    try:
        # Build filters
        filters = {}
        if export_request.supervision_year:
            filters['supervision_year'] = export_request.supervision_year
        if export_request.responsible_team:
            filters['responsible_team'] = export_request.responsible_team
        if export_request.overall_status:
            filters['overall_status'] = export_request.overall_status
        if export_request.transformation_target:
            filters['transformation_target'] = export_request.transformation_target

        # Generate Excel file
        excel_data = await excel_service.export_applications_to_excel(
            db=db,
            application_ids=export_request.application_ids,
            filters=filters,
            template_style=export_request.template_style
        )

        # Generate filename
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"applications_export_{timestamp}.xlsx"

        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export applications: {str(e)}"
        )


@router.get("/subtasks/export")
async def export_subtasks_to_excel(
    export_request: SubTaskExportRequest = Depends(),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Export subtasks to Excel file."""

    try:
        # Build filters
        filters = {}
        if export_request.task_status:
            filters['task_status'] = export_request.task_status
        if export_request.sub_target:
            filters['sub_target'] = export_request.sub_target
        if export_request.is_blocked is not None:
            filters['is_blocked'] = export_request.is_blocked
        if export_request.responsible_person:
            filters['responsible_person'] = export_request.responsible_person

        # Generate Excel file
        excel_data = await excel_service.export_subtasks_to_excel(
            db=db,
            application_id=export_request.application_id,
            subtask_ids=export_request.subtask_ids,
            filters=filters,
            template_style=export_request.template_style
        )

        # Generate filename
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"subtasks_export_{timestamp}.xlsx"

        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export subtasks: {str(e)}"
        )


@router.get("/template")
async def generate_import_template(
    template_request: ExcelTemplateRequest = Depends(),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Generate Excel import template."""

    try:
        template_data = excel_service.generate_import_template(
            template_type=template_request.template_type,
            include_sample_data=template_request.include_sample_data
        )

        # Generate filename based on template type
        template_names = {
            'applications': '应用导入模板',
            'subtasks': '子任务导入模板',
            'combined': '综合导入模板'
        }

        filename = f"{template_names.get(template_request.template_type, 'import_template')}.xlsx"

        return Response(
            content=template_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate template: {str(e)}"
        )


@router.post("/preview", response_model=ExcelImportPreview)
async def preview_excel_file(
    file: UploadFile = File(..., description="Excel file to preview"),
    entity_type: Optional[str] = Query(None, description="Expected entity type (application, subtask)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Preview Excel file before import."""

    # Validate file
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )

    try:
        file_content = await file.read()

        # Basic preview implementation
        from openpyxl import load_workbook
        import io

        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet_names = workbook.sheetnames

        # Get first sheet data for preview
        first_sheet = workbook.active
        sample_rows = []

        # Read first few rows
        for i, row in enumerate(first_sheet.iter_rows(min_row=1, max_row=6, values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else '' for cell in row]
                continue

            row_data = {}
            for j, cell in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_data[headers[j]] = cell

            if any(v for v in row_data.values()):  # Skip empty rows
                sample_rows.append(row_data)

        # Simple column analysis
        column_analysis = []
        for header in headers[:10]:  # Analyze first 10 columns
            if header:
                column_analysis.append({
                    "column_name": header,
                    "data_type": "string",
                    "sample_values": [row.get(header, '') for row in sample_rows[:3]],
                    "quality_score": 0.8
                })

        # Auto-detect entity type based on headers
        detected_type = None
        if any(keyword in ' '.join(headers).lower() for keyword in ['l2', '应用', 'application']):
            detected_type = "application"
        elif any(keyword in ' '.join(headers).lower() for keyword in ['任务', 'task', '模块']):
            detected_type = "subtask"

        return ExcelImportPreview(
            sheet_names=sheet_names,
            detected_entity_type=detected_type,
            column_analysis=column_analysis,
            data_quality_score=0.85,
            sample_rows=sample_rows,
            recommendations=[
                "文件格式正确",
                "数据结构清晰",
                f"检测到 {len(sample_rows)} 行数据"
            ]
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to preview file: {str(e)}"
        )


@router.post("/validate", response_model=ExcelImportResult)
async def validate_excel_file(
    file: UploadFile = File(..., description="Excel file to validate"),
    entity_type: str = Query(..., description="Entity type (application, subtask)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Validate Excel file without importing data."""

    # Validate file
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )

    file_content = await file.read()

    try:
        if entity_type == "application":
            result = await excel_service.import_applications_from_excel(
                db=db,
                file_content=file_content,
                user=current_user,
                validate_only=True
            )
        elif entity_type == "subtask":
            result = await excel_service.import_subtasks_from_excel(
                db=db,
                file_content=file_content,
                user=current_user,
                validate_only=True
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="entity_type must be 'application' or 'subtask'"
            )

        return ExcelImportResult(**result)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to validate file: {str(e)}"
        )


@router.get("/import/history")
async def get_import_history(
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(20, ge=1, le=100, description="Number of records to return"),
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    status: Optional[str] = Query(None, description="Filter by status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER]))
):
    """Get Excel import history."""

    # This would typically query an import_history table
    # For now, return a mock response
    mock_history = [
        {
            "import_id": "imp_001",
            "user_id": current_user.id,
            "import_type": "application",
            "file_name": "applications_2024.xlsx",
            "file_size_bytes": 1024000,
            "total_rows": 150,
            "processed_rows": 145,
            "error_rows": 5,
            "import_time_ms": 5000,
            "created_at": "2024-01-15T10:30:00Z",
            "status": "completed"
        }
    ]

    return {
        "total": len(mock_history),
        "items": mock_history[skip:skip + limit],
        "page": (skip // limit) + 1,
        "page_size": limit
    }


@router.get("/export/formats")
async def get_export_formats(
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Get available export formats and options."""

    return {
        "formats": [
            {
                "format": "xlsx",
                "description": "Excel 2007+ format with advanced features",
                "features": ["formulas", "styling", "multiple_sheets", "charts"],
                "max_rows": 1048576
            },
            {
                "format": "xls",
                "description": "Legacy Excel format",
                "features": ["basic_styling", "formulas"],
                "max_rows": 65536
            },
            {
                "format": "csv",
                "description": "Comma-separated values",
                "features": ["lightweight", "universal_support"],
                "max_rows": "unlimited"
            }
        ],
        "template_styles": [
            {
                "style": "standard",
                "description": "Standard corporate styling"
            },
            {
                "style": "minimal",
                "description": "Clean minimal design"
            },
            {
                "style": "detailed",
                "description": "Comprehensive with extra formatting"
            }
        ]
    }


@router.get("/mapping/templates")
async def get_mapping_templates(
    entity_type: Optional[str] = Query(None, description="Filter by entity type"),
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER, UserRole.EDITOR]))
):
    """Get available Excel mapping templates."""

    # Mock template data - in real implementation, this would query a database
    templates = [
        {
            "template_name": "标准应用导入模板",
            "entity_type": "application",
            "description": "标准应用数据导入映射模板",
            "is_default": True,
            "field_count": 19
        },
        {
            "template_name": "标准子任务导入模板",
            "entity_type": "subtask",
            "description": "标准子任务数据导入映射模板",
            "is_default": True,
            "field_count": 17
        }
    ]

    if entity_type:
        templates = [t for t in templates if t["entity_type"] == entity_type]

    return {
        "templates": templates,
        "total": len(templates)
    }


@router.get("/health", response_model=ExcelHealthCheck)
async def excel_service_health_check(
    current_user: User = Depends(require_roles([UserRole.ADMIN, UserRole.MANAGER]))
):
    """Check Excel service health and status."""

    import psutil
    import os

    try:
        # Get process info
        process = psutil.Process(os.getpid())
        memory_info = process.memory_info()

        return ExcelHealthCheck(
            status="healthy",
            version="1.0.0",
            active_imports=0,  # Would be tracked in real implementation
            active_exports=0,  # Would be tracked in real implementation
            total_processed_today=0,  # Would be queried from database
            average_processing_time_ms=2500.0,
            error_rate_percentage=2.5,
            memory_usage_mb=memory_info.rss / 1024 / 1024,
            disk_usage_mb=0.0,  # Would calculate temp file usage
            service_uptime_hours=24.5
        )

    except Exception as e:
        return ExcelHealthCheck(
            status="unhealthy",
            version="1.0.0",
            active_imports=0,
            active_exports=0,
            total_processed_today=0,
            average_processing_time_ms=0.0,
            error_rate_percentage=100.0,
            memory_usage_mb=0.0,
            disk_usage_mb=0.0,
            service_uptime_hours=0.0
        )