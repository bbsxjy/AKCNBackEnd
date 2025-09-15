"""
Test script for audit log export functionality
"""

import asyncio
import aiohttp
import json
from datetime import datetime

BASE_URL = "http://localhost:8000/api/v1"


async def test_export_json():
    """Test export to JSON format."""

    async with aiohttp.ClientSession() as session:
        headers = {
            "Authorization": "Bearer YOUR_TOKEN_HERE",  # Replace with actual token
            "Content-Type": "application/json"
        }

        # Test JSON export
        print("Testing JSON export...")
        export_data = {
            "format": "json",
            "table_name": "applications",
            "start_date": "2024-01-01",
            "end_date": "2024-12-31"
        }

        try:
            async with session.post(
                f"{BASE_URL}/audit/export",
                headers=headers,
                json=export_data
            ) as response:
                if response.status == 200:
                    result = await response.json()
                    print(f"✅ JSON export successful!")
                    print(f"  Total records: {result['total_records']}")
                    print(f"  Export format: {result['export_format']}")
                    print(f"  Filters applied: {json.dumps(result['filters_applied'], indent=2)}")
                    if result['data']:
                        print(f"  Sample record: {json.dumps(result['data'][0], indent=2)[:200]}...")
                else:
                    error = await response.text()
                    print(f"❌ JSON export failed: {response.status}")
                    print(f"  Error: {error}")
        except Exception as e:
            print(f"❌ Error during JSON export: {e}")


async def test_export_csv():
    """Test export to CSV format."""

    async with aiohttp.ClientSession() as session:
        headers = {
            "Authorization": "Bearer YOUR_TOKEN_HERE",  # Replace with actual token
            "Content-Type": "application/json"
        }

        # Test CSV export
        print("\nTesting CSV export...")
        export_data = {
            "format": "csv",
            "table_name": "applications",
            "operation": "UPDATE"
        }

        try:
            async with session.post(
                f"{BASE_URL}/audit/export",
                headers=headers,
                json=export_data
            ) as response:
                if response.status == 200:
                    # Save CSV file
                    content = await response.read()
                    filename = f"audit_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

                    with open(filename, 'wb') as f:
                        f.write(content)

                    print(f"✅ CSV export successful!")
                    print(f"  File saved as: {filename}")
                    print(f"  File size: {len(content)} bytes")

                    # Show first few lines
                    lines = content.decode('utf-8-sig').split('\n')[:3]
                    print(f"  Preview:")
                    for line in lines:
                        print(f"    {line[:100]}...")
                else:
                    error = await response.text()
                    print(f"❌ CSV export failed: {response.status}")
                    print(f"  Error: {error}")
        except Exception as e:
            print(f"❌ Error during CSV export: {e}")


async def test_export_excel():
    """Test export to Excel format."""

    async with aiohttp.ClientSession() as session:
        headers = {
            "Authorization": "Bearer YOUR_TOKEN_HERE",  # Replace with actual token
            "Content-Type": "application/json"
        }

        # Test Excel export
        print("\nTesting Excel export...")
        export_data = {
            "format": "excel",
            "start_date": "2024-01-01",
            "end_date": "2024-12-31"
        }

        try:
            async with session.post(
                f"{BASE_URL}/audit/export",
                headers=headers,
                json=export_data
            ) as response:
                if response.status == 200:
                    # Save Excel file
                    content = await response.read()
                    filename = f"audit_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                    with open(filename, 'wb') as f:
                        f.write(content)

                    print(f"✅ Excel export successful!")
                    print(f"  File saved as: {filename}")
                    print(f"  File size: {len(content)} bytes")

                    # Try to open with openpyxl to verify
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(filename=filename)
                        print(f"  Sheets: {wb.sheetnames}")
                        ws = wb.active
                        print(f"  Rows in main sheet: {ws.max_row}")
                        print(f"  Columns in main sheet: {ws.max_column}")
                    except ImportError:
                        print("  (Install openpyxl to verify Excel file structure)")
                else:
                    error = await response.text()
                    print(f"❌ Excel export failed: {response.status}")
                    print(f"  Error: {error}")
        except Exception as e:
            print(f"❌ Error during Excel export: {e}")


async def test_export_with_filters():
    """Test export with various filters."""

    async with aiohttp.ClientSession() as session:
        headers = {
            "Authorization": "Bearer YOUR_TOKEN_HERE",  # Replace with actual token
            "Content-Type": "application/json"
        }

        # Test with multiple filters
        print("\nTesting export with filters...")
        export_data = {
            "format": "json",
            "table_name": "applications",
            "operation": "UPDATE",
            "user_id": 1,
            "start_date": "2024-01-01",
            "end_date": "2024-12-31",
            "filters": {
                "custom_filter": "test"
            }
        }

        try:
            async with session.post(
                f"{BASE_URL}/audit/export",
                headers=headers,
                json=export_data
            ) as response:
                if response.status == 200:
                    result = await response.json()
                    print(f"✅ Filtered export successful!")
                    print(f"  Total records after filtering: {result['total_records']}")
                    print(f"  Applied filters:")
                    for key, value in result['filters_applied'].items():
                        if value:
                            print(f"    - {key}: {value}")
                else:
                    error = await response.text()
                    print(f"❌ Filtered export failed: {response.status}")
                    print(f"  Error: {error}")
        except Exception as e:
            print(f"❌ Error during filtered export: {e}")


async def test_invalid_format():
    """Test with invalid format."""

    async with aiohttp.ClientSession() as session:
        headers = {
            "Authorization": "Bearer YOUR_TOKEN_HERE",  # Replace with actual token
            "Content-Type": "application/json"
        }

        # Test with invalid format
        print("\nTesting invalid format...")
        export_data = {
            "format": "invalid_format"
        }

        try:
            async with session.post(
                f"{BASE_URL}/audit/export",
                headers=headers,
                json=export_data
            ) as response:
                if response.status == 200:
                    # Should default to JSON
                    result = await response.json()
                    print(f"✅ Invalid format defaulted to JSON")
                    print(f"  Export format: {result['export_format']}")
                else:
                    error = await response.text()
                    print(f"  Response: {response.status}")
                    print(f"  Details: {error}")
        except Exception as e:
            print(f"❌ Error: {e}")


if __name__ == "__main__":
    print("=== Audit Log Export Test Suite ===")
    print(f"Started at: {datetime.now()}")

    # Run all tests
    asyncio.run(test_export_json())
    asyncio.run(test_export_csv())
    asyncio.run(test_export_excel())
    asyncio.run(test_export_with_filters())
    asyncio.run(test_invalid_format())

    print(f"\nCompleted at: {datetime.now()}")
    print("=== Test Suite Complete ===")