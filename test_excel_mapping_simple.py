"""
Simple test script to verify Excel column mapping is working correctly
"""
# -*- coding: utf-8 -*-
import sys
import io

# Set output encoding to UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from app.services.excel_service import ExcelService, ExcelMappingConfig

# Create a test Excel file with exact column headers as specified
wb = Workbook()
ws = wb.active
ws.title = "应用总追踪表"

# Set headers exactly as specified by user
headers = [
    "L2ID",
    "L2应用",
    "监管验收年份",
    "改造目标",
    "是否已完成AK",
    "是否已完成云原生",
    "当前改造阶段",
    "改造状态",
    "【计划】\n需求完成时间",
    "【计划】\n发版时间",
    "【计划】\n技术上线时间",
    "【计划】\n业务上线时间",
    "【实际】\n需求到达时间",
    "【实际】\n发版时间",
    "【实际】\n技术上线时间",
    "【实际】\n业务上线时间",
    "备注",
    "档位",
    "所属L1",
    "所属项目",
    "开发模式",
    "运维模式",
    "开发负责人",
    "开发团队",
    "运维负责人",
    "运维团队",
    "所属指标",
    "验收状态"
]

# Write headers
for i, header in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=header)

# Add sample data
sample_data = [
    ["10001", "测试应用1", 2024, "AK", "否", "否", "开发中", "研发进行中",
     "2024-03-01", "2024-03-15", "2024-04-01", "2024-04-15",
     "2024-03-05", "2024-03-20", "2024-04-10", "2024-04-20",
     "测试备注", 1, "数字金融", "AK改造项目", "敏捷", "DevOps",
     "张三", "团队A", "李四", "运维团队B", "KPI-001", "待验收"],
    ["10002", "测试应用2", 2024, "云原生", "是", "否", "测试中", "业务上线中",
     "2024-02-01", "2024-02-15", "2024-03-01", "2024-03-15",
     "2024-02-05", "2024-02-20", "2024-03-10", "2024-03-20",
     "另一个测试", 2, "零售金融", "云原生改造", "瀑布", "传统运维",
     "王五", "团队B", "赵六", "运维团队A", "KPI-002", "已验收"]
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Save to bytes for testing
from openpyxl import load_workbook

print("="*60)
print("Testing Excel column mapping")
print("="*60)

# Test the mapping directly
service = ExcelService()
config = ExcelMappingConfig()

print("\n1. Testing header mapping:")
print("-"*40)

# Test each header to see if it maps correctly
for header in headers:
    if header in config.APPLICATION_FIELDS:
        mapped_field = config.APPLICATION_FIELDS[header]
        print(f"[MAPPED] '{header}' -> '{mapped_field}'")
    else:
        print(f"[NOT MAPPED] '{header}'")

# Now test with the actual worksheet_to_dataframe method
print("\n2. Testing worksheet_to_dataframe method:")
print("-"*40)

# Load the workbook
df = service._worksheet_to_dataframe(ws, config.APPLICATION_FIELDS)

print(f"DataFrame shape: {df.shape}")
print(f"DataFrame columns: {list(df.columns)}")

if len(df) > 0:
    print("\n3. Sample data from DataFrame:")
    print("-"*40)
    for idx, row in df.head(2).iterrows():
        print(f"\nRow {idx + 1}:")
        for col in ['l2_id', 'app_name', 'dev_team', 'overall_transformation_target', 'current_status']:
            if col in row:
                print(f"  {col}: {row[col]}")

    # Check specifically for app_name
    if 'app_name' in df.columns:
        print(f"\n4. app_name column analysis:")
        print("-"*40)
        print(f"Total rows: {len(df)}")
        non_empty = df[df['app_name'].notna() & (df['app_name'] != '')]
        print(f"Rows with non-empty app_name: {len(non_empty)}")
        if len(non_empty) > 0:
            print(f"app_name values: {list(non_empty['app_name'])}")
    else:
        print("\n[WARNING]: app_name column not found in DataFrame!")

print("\n" + "="*60)
print("Test completed!")
print("="*60)